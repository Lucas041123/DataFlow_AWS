import sys
import os
import glob
import polars as pl
import openpyxl 
import xlrd
import json
import re
from unidecode import unidecode
from collections import defaultdict, Counter

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLineEdit, QLabel, QListWidget, QListWidgetItem,
    QComboBox, QProgressBar, QTextEdit, QFileDialog, QTabWidget, 
    QTableView, QDialogButtonBox, QTableWidget, QDialog, QTableWidgetItem,
    QCheckBox, QHeaderView, QScrollArea, QGroupBox, QAbstractItemView, QStyle,
    QInputDialog
)
from PySide6.QtCore import Qt, QThread, Signal , QAbstractTableModel
from PySide6.QtGui import QColor, QPalette, QIcon, QAction, QTextCursor
from PySide6.QtSvgWidgets import QSvgWidget
from PySide6.QtWidgets import QRadioButton
from enum import Enum

# Definir os tipos de dados que o usuário pode escolher
DATA_TYPES_OPTIONS = ["Automático/String", "Inteiro", "Decimal (Float)", "Data", "Booleano"]
# Mapeamento para tipos Polars (pode ser um dict global ou dentro do worker)
TYPE_STRING_TO_POLARS = {
    "Automático/String": pl.String,
    "Inteiro": pl.Int64,
    "Decimal (Float)": pl.Float64,
    "Data": pl.Date, # ou pl.Datetime se precisar de hora
    "Booleano": pl.Boolean
}
# --- Constantes para as Opções de Filtro ---
OPERATOR_OPTIONS = [
    "Igual a",
    "Diferente de",
    "Contém",
    "Não contém",
    "Começa com",
    "Termina com",
    "Maior que",
    "Menor que",
    "Entre",
    "Está em branco",
    "Não está em branco",
]
# Operadores que não precisam de um campo de valor
OPERATORS_NO_VALUE = {"Está em branco", "Não está em branco"}

CONFIG_FILE_NAME = "config_consolidador.json" # Nome do arquivo de configuração

def _normalize_header_name(header_name: str) -> str:
    if not isinstance(header_name, str):
        header_name = str(header_name)
    
    # 1. Remove acentos (ex: "Endereço" -> "Endereco")
    text = unidecode(header_name)
    # 2. Converte para minúsculas
    text = text.lower()
    # 3. Substitui separadores comuns por espaço (para depois lidar com "valorICMS" vs "valor_ICMS")
    text = re.sub(r'[._-]+', ' ', text)
    # 4. Adiciona espaço antes de letras maiúsculas (camelCase -> camel Case)
    text = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', text)
    # 5. Remove todos os caracteres não alfanuméricos e junta tudo
    text = re.sub(r'[^a-z0-9]', '', text)
    return text

def _find_header_row_index(df_sample: pl.DataFrame, max_rows_to_check: int = 20) -> int:
    """
    Analisa as primeiras N linhas de um DataFrame e retorna o índice da linha
    que tem a maior probabilidade de ser o cabeçalho, usando uma heurística de
    transição de tipos.
    """
    best_score = -999
    header_row_index = 0
    rows_to_check = min(df_sample.height, max_rows_to_check)

    # Nós só podemos checar até a penúltima linha, pois sempre olhamos para a linha seguinte
    for i in range(rows_to_check - 1):
        row_current = df_sample.row(i)
        row_next = df_sample.row(i + 1)
        
        # --- Cálculo da Pontuação da Linha Atual (baseado na lógica anterior) ---
        base_score = 0
        string_count = 0
        null_count = 0
        numeric_string_count = 0
        
        for value in row_current:
            if value is None: null_count += 1
            elif isinstance(value, str) and value.strip():
                string_count += 1
                if value.strip().isnumeric(): numeric_string_count += 1
        
        num_cells = len(row_current)
        non_null_cells = num_cells - null_count

        if non_null_cells == 0:
            base_score = -100
        else:
            uniqueness_ratio = len(set(v for v in row_current if v is not None)) / non_null_cells
            string_ratio = string_count / non_null_cells
            numeric_string_ratio = numeric_string_count / non_null_cells
            base_score += uniqueness_ratio * 3
            base_score += string_ratio * 3
            base_score -= numeric_string_ratio * 5
            base_score -= (null_count / num_cells) * 2

        # --- NOVA HEURÍSTICA: Pontuação por Transição de Tipos ---
        transition_score = 0
        for j in range(num_cells):
            type_current = type(row_current[j])
            type_next = type(row_next[j])
            
            # Recompensa fortemente a transição de Texto para Número
            if type_current is str and type_next in (int, float):
                transition_score += 1
            # Penaliza levemente se os tipos são iguais (exceto None)
            elif type_current is not None and type_current == type_next:
                transition_score -= 0.2

        transition_bonus = (transition_score / num_cells) * 5 # Bônus forte
        
        # Combina as pontuações
        final_score = base_score + transition_bonus
        
        if final_score > best_score:
            best_score = final_score
            header_row_index = i
            
    return header_row_index

def _make_headers_unique(header_names: list) -> list:
    """
    Garante que todos os nomes de cabeçalho em uma lista sejam únicos
    """
    counts = Counter(header_names)
    duplicates = {name for name, count in counts.items() if count > 1}
    if not duplicates:
            return header_names
    new_headers = []
    running_counts = Counter()
    for name in header_names:
        if name in duplicates:
            running_counts[name] += 1
            new_headers.append(f"{name}_{running_counts[name]}")
        else:
            new_headers.append(name)
    return new_headers

class LogLevel(Enum):
    INFO = "[INFO]"
    WARNING = "[AVISO]"
    ERROR = "[ERRO]"
    SUCCESS = "[SUCESSO]"

class PivotDialog(QDialog):
    """Um diálogo para configurar a operação de tabela dinâmica (pivot)."""
    def __init__(self, all_headers, numeric_headers, existing_rules=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configurar Tabela de Resumo (Dinâmica)")
        self.setMinimumSize(750, 550) # Ajuste de tamanho

        self.all_headers = all_headers
        self.numeric_headers = numeric_headers
        self.operations = ["Soma", "Média", "Contagem", "Mínimo", "Máximo", "Contagem Única"]

        # Layout Principal
        main_layout = QVBoxLayout(self)

        # --- SEÇÃO DE AGRUPAMENTO (COM A MELHORIA DE UI) ---
        group_by_box = QGroupBox("1. Agrupar por (Linhas)")
        group_by_v_layout = QVBoxLayout(group_by_box)
        group_by_v_layout.addWidget(QLabel("Selecione as colunas para agrupar (esquerda) e veja sua seleção (direita):"))

        # Layout horizontal para as duas listas
        group_by_h_layout = QHBoxLayout()

        # Painel da Esquerda: Lista de todas as colunas para seleção
        self.group_by_list = QListWidget()
        self.group_by_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.group_by_list.addItems(self.all_headers)
        # Conecta a mudança na seleção à atualização do painel da direita
        self.group_by_list.itemSelectionChanged.connect(self.update_group_by_display)
        group_by_h_layout.addWidget(self.group_by_list)

        # Painel da Direita: Apenas para exibir as colunas selecionadas
        self.selected_group_by_display = QListWidget()
        self.selected_group_by_display.setMaximumWidth(220) # Largura fixa para o painel de exibição
        # Desabilita a interação direta com este painel
        self.selected_group_by_display.setSelectionMode(QAbstractItemView.NoSelection)
        self.selected_group_by_display.setStyleSheet("QListWidget { background-color: #ECECEC; }") # Cor de fundo para indicar que é um display
        group_by_h_layout.addWidget(self.selected_group_by_display)

        group_by_v_layout.addLayout(group_by_h_layout)
        main_layout.addWidget(group_by_box)

        # --- Seção de Agregações (sem alterações) ---
        agg_box = QGroupBox("2. Cálculos (Valores)")
        agg_layout = QVBoxLayout(agg_box)
        
        agg_actions_layout = QHBoxLayout()
        agg_actions_layout.addWidget(QLabel("Defina as colunas e as operações de cálculo:"))
        agg_actions_layout.addStretch()
        self.add_agg_button = QPushButton("Adicionar Cálculo")
        self.add_agg_button.clicked.connect(self.add_aggregation_row)
        agg_actions_layout.addWidget(self.add_agg_button)
        agg_layout.addLayout(agg_actions_layout)

        self.agg_table = QTableWidget()
        self.agg_table.setColumnCount(3)
        self.agg_table.setHorizontalHeaderLabels(["Coluna para Calcular", "Operação", "Ação"])
        self.agg_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.agg_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.agg_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        agg_layout.addWidget(self.agg_table)
        main_layout.addWidget(agg_box)

        self.only_pivot_checkbox = QCheckBox("Gerar apenas a Tabela de Resumo (ignorar dados consolidados na saída)")
        main_layout.addWidget(self.only_pivot_checkbox)

        # --- Botões Finais (sem alterações) ---
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel | QDialogButtonBox.Reset)
        self.button_box.button(QDialogButtonBox.Reset).setText("Limpar Regras")
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.button_box.button(QDialogButtonBox.Reset).clicked.connect(self.clear_rules)
        main_layout.addWidget(self.button_box)

        if existing_rules:
            self.populate_from_rules(existing_rules)
        else:
            self.add_aggregation_row()
    
    def update_group_by_display(self):
        """NOVO: Atualiza o painel da direita para espelhar a seleção de agrupamento."""
        self.selected_group_by_display.clear()
        selected_items = self.group_by_list.selectedItems()
        if selected_items:
            # Pega o texto de cada item selecionado e adiciona ao painel de exibição
            self.selected_group_by_display.addItems(sorted([item.text() for item in selected_items]))

    def add_aggregation_row(self, rule=None):
        """Adiciona uma linha à tabela de agregação."""
        row_position = self.agg_table.rowCount()
        self.agg_table.insertRow(row_position)

        # Coluna 0: Coluna de Valor
        col_combo = QComboBox()
        col_combo.addItems(self.numeric_headers)
        self.agg_table.setCellWidget(row_position, 0, col_combo)

        # Coluna 1: Operação
        op_combo = QComboBox()
        op_combo.addItems(self.operations)
        self.agg_table.setCellWidget(row_position, 1, op_combo)
        
        # Coluna 2: Botão de Remover
        remove_button = QPushButton("X")
        remove_button.setFixedWidth(40)
        # Usamos uma lambda para passar a linha correta no momento do clique
        remove_button.clicked.connect(lambda: self.agg_table.removeRow(self.agg_table.indexAt(remove_button.pos()).row()))
        self.agg_table.setCellWidget(row_position, 2, remove_button)
        
        # Preenche com dados da regra, se fornecida
        if rule:
            col_combo.setCurrentText(rule.get("column", ""))
            op_combo.setCurrentText(rule.get("operation", ""))

    def populate_from_rules(self, rules):
        """Preenche o diálogo com regras existentes."""
        # Limpa o que já existe
        self.group_by_list.clearSelection()
        while self.agg_table.rowCount() > 0:
            self.agg_table.removeRow(0)

        # Popula agrupamentos
        for item_text in rules.get("group_by", []):
            items = self.group_by_list.findItems(item_text, Qt.MatchExactly)
            if items:
                items[0].setSelected(True)
        
        self.update_group_by_display()
        # Popula agregações
        for rule in rules.get("aggregations", []):
            self.add_aggregation_row(rule)
        self.only_pivot_checkbox.setChecked(rules.get("only_pivot", False))
        

    def clear_rules(self):
        """Limpa todas as seleções e regras no diálogo."""
        self.group_by_list.clearSelection()
        while self.agg_table.rowCount() > 0:
            self.agg_table.removeRow(0)
        self.add_aggregation_row() # Adiciona uma linha em branco
        
    def get_rules(self):
        """Lê os widgets e retorna um dicionário com as regras de pivot."""
        group_by_cols = [item.text() for item in self.group_by_list.selectedItems()]
        
        aggregations = []
        for row in range(self.agg_table.rowCount()):
            col_combo = self.agg_table.cellWidget(row, 0)
            op_combo = self.agg_table.cellWidget(row, 1)
            
            if col_combo and op_combo and col_combo.currentText():
                aggregations.append({
                    "column": col_combo.currentText(),
                    "operation": op_combo.currentText()
                })
        
        # Só retorna regras se o usuário definiu agrupamentos e agregações
        if not group_by_cols or not aggregations:
            return {}

        return {
            "group_by": group_by_cols,
            "aggregations": aggregations,
            "only_pivot": self.only_pivot_checkbox.isChecked() 
        }

class FilterDialog(QDialog):
    def __init__(self, final_headers, existing_filters=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Definir Filtros de Dados")
        self.setMinimumSize(600, 400)

        self.final_headers = sorted(final_headers)
        self.filter_rows = []  # Para manter referência aos widgets de cada linha

        # --- Layout Principal ---
        main_layout = QVBoxLayout(self)

        # --- Descrição ---
        description_label = QLabel("Adicione regras para filtrar os dados. Regras na mesma coluna (exceto 'Entre') são unidas com OU. Grupos de colunas diferentes são unidos com E.")
        main_layout.addWidget(description_label)

        # --- Layout para as Linhas de Filtro ---
        self.filters_layout = QVBoxLayout()
        self.filters_layout.setAlignment(Qt.AlignTop)

        # Widget e ScrollArea para conter as linhas de filtro
        filters_widget = QWidget()
        filters_widget.setLayout(self.filters_layout)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(filters_widget)
        main_layout.addWidget(scroll_area)

        # --- Botões de Ação ---
        add_filter_button = QPushButton("Adicionar Filtro")
        add_filter_button.clicked.connect(self.add_filter_row)
        main_layout.addWidget(add_filter_button)

        # Botões OK e Cancelar
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        main_layout.addWidget(self.button_box)

        # Preencher com filtros existentes ou iniciar com uma linha em branco
        if existing_filters:
            for f in existing_filters:
                self.add_filter_row(filter_data=f)
        else:
            self.add_filter_row()

    def add_filter_row(self, filter_data=None):
        """Adiciona uma nova linha de widget para criar uma regra de filtro."""
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)

        column_combo = QComboBox()
        if self.final_headers:
            column_combo.addItems(self.final_headers)
        else:
            column_combo.addItem("Nenhum cabeçalho mapeado"); column_combo.setEnabled(False)

        operator_combo = QComboBox()
        operator_combo.addItems(OPERATOR_OPTIONS)

        value1_edit = QLineEdit()
        value1_edit.setPlaceholderText("Valor")
        and_label = QLabel(" e ")
        value2_edit = QLineEdit()
        value2_edit.setPlaceholderText("Valor Máximo")
        
        remove_button = QPushButton("X"); remove_button.setFixedWidth(30)
        
        row_layout.addWidget(column_combo)
        row_layout.addWidget(operator_combo)
        row_layout.addWidget(value1_edit)
        row_layout.addWidget(and_label)
        row_layout.addWidget(value2_edit)
        row_layout.addWidget(remove_button)

        # Guarda as referências aos widgets da linha para fácil acesso
        row_widgets = {"widget": row_widget, "op_combo": operator_combo, "val1": value1_edit, "and_label": and_label, "val2": value2_edit}
        self.filter_rows.append(row_widgets)

        self.filters_layout.addWidget(row_widget)

        # Conecta o botão de remover
        remove_button.clicked.connect(lambda: self.remove_filter_row(row_widgets))
        # Conecta a mudança do operador à lógica de visibilidade
        operator_combo.currentTextChanged.connect(lambda text: self._on_operator_changed(text, row_widgets))
        
        # Preenche os dados se for um filtro existente
        if filter_data:
            column_combo.setCurrentText(filter_data.get("column", ""))
            operator_combo.setCurrentText(filter_data.get("operator", ""))
            if isinstance(filter_data.get("value"), list): # Se for "Entre"
                value1_edit.setText(filter_data["value"][0])
                value2_edit.setText(filter_data["value"][1])
            else:
                value1_edit.setText(filter_data.get("value", ""))

        # Dispara a lógica de visibilidade para o estado inicial
        self._on_operator_changed(operator_combo.currentText(), row_widgets)

    def _on_operator_changed(self, text, row_widgets):
        """Ajusta a visibilidade dos campos de valor com base no operador."""
        is_between = (text == "Entre")
        is_no_value = (text in OPERATORS_NO_VALUE)
        
        row_widgets["val1"].setVisible(not is_no_value)
        row_widgets["and_label"].setVisible(is_between)
        row_widgets["val2"].setVisible(is_between)

    def remove_filter_row(self, row_widget):
        """Remove uma linha de filtro da interface e da nossa lista de referência."""
        if row_widget in self.filter_rows:
            row_widget['widget'].deleteLater()
            self.filter_rows.remove(row_widget)

    def get_filters(self):
        """Lê todos os widgets de filtro e retorna uma lista de dicionários com as regras."""
        rules = []
        for row_data in self.filter_rows:
            # Encontrar os widgets filhos pelo nome do objeto seria mais robusto, mas isso funciona
            column_combo = row_data["widget"].findChild(QComboBox)
            operator = row_data["op_combo"].currentText()
            
            value = ""
            if operator == "Entre":
                val1 = row_data["val1"].text()
                val2 = row_data["val2"].text()
                value = [val1, val2] # Salva como uma lista
            else:
                value = row_data["val1"].text()

            rules.append({
                "column": column_combo.currentText(),
                "operator": operator,
                "value": value,
            })
        return rules

class ConsolidationWorker(QThread):
    progress_updated = Signal(int) 
    log_message = Signal(str, LogLevel) 
    finished = Signal(bool, str) 
    progress_text_updated = Signal(str)

    def __init__(self, files_to_process, output_path, output_format, header_mapping, filter_rules, delimiter, pivot_rules, duplicates_config=None):
        super().__init__()
        self.files_to_process = files_to_process 
        self.output_path = output_path
        self.output_format = output_format
        self.header_mapping = header_mapping
        self.filter_rules = filter_rules
        self.pivot_rules = pivot_rules
        self.duplicates_config = duplicates_config or {}
        self.delimiter = delimiter
        self.is_running = True

    def run(self):
        try:
            self.log_message.emit("Iniciando processo de consolidação...", LogLevel.INFO)
            all_dataframes_processed = [] 
            
            total_items = 0
            for _, sheets_to_process_for_file in self.files_to_process:
                if sheets_to_process_for_file is None: total_items += 1
                else: total_items += len(sheets_to_process_for_file)
            if total_items == 0:
                self.log_message.emit("Nenhum item válido para processar.", LogLevel.WARNING)
                self.finished.emit(False, "Nenhum item para processar.")
                return
            processed_items = 0

            for file_path, selected_sheets in self.files_to_process:
                if not self.is_running: break 
                file_name = os.path.basename(file_path)
                sheets_to_iterate = selected_sheets if selected_sheets is not None else [None]

                for sheet_name in sheets_to_iterate:
                    if not self.is_running: break 
                    current_item_description = f"'{file_name}'" + (f" - Aba: '{sheet_name}'" if sheet_name else "")
                    try:
                        # ETAPA 1: Pré-leitura e Detecção do Cabeçalho
                        n_preread_rows = 20
                        header_row_index = 0
                        header_names = []
                        
                        pre_read_df = None
                        if file_path.lower().endswith((".csv", ".txt")):
                            pre_read_df = pl.read_csv(source=file_path, has_header=False, n_rows=n_preread_rows, separator=self.delimiter, encoding='latin-1', ignore_errors=True, infer_schema = False, quote_char = None, truncate_ragged_lines = True)
                        elif file_path.lower().endswith((".xlsx", ".xls")):
                            pre_read_df = pl.read_excel(source=file_path, sheet_name=sheet_name, has_header = False).head(n_preread_rows)

                        if pre_read_df is not None and not pre_read_df.is_empty():
                            header_row_index = _find_header_row_index(pre_read_df, n_preread_rows)
                            header_names_raw = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(pre_read_df.row(header_row_index))]
                            header_names = _make_headers_unique(header_names_raw)
                        
                        # --- LÓGICA DE LEITURA FINAL E ROBUSTA ---
                        df_original = None
                        
                        # Ler o arquivo inteiro como dados brutos, sem cabeçalho
                        if file_path.lower().endswith((".csv", ".txt")):
                            df_raw_data = pl.read_csv(source=file_path, has_header=False, separator=self.delimiter, encoding='latin-1', infer_schema = False, ignore_errors = True, quote_char = None, truncate_ragged_lines = True)
                        elif file_path.lower().endswith((".xlsx", ".xls")):
                            df_raw_data = pl.read_excel(source=file_path, sheet_name=sheet_name, has_header = False)
                        
                        if df_raw_data is not None and not df_raw_data.is_empty():
                            # Fatiar o DataFrame para remover lixo + linha do cabeçalho
                            df_data_only = df_raw_data.slice(offset=header_row_index + 1)
                            
                            if not df_data_only.is_empty():
                                # Renomear as colunas com os nomes que detectamos
                                rename_mapping = {old_name: new_name for old_name, new_name in zip(df_data_only.columns, header_names)}
                                df_original = df_data_only.rename(rename_mapping)

                        # Sua lógica de fallback para erros de conversão no Excel pode ser integrada aqui se necessário,
                        # mas esta leitura primária é muito mais estável.
                        
                        if df_original is None or df_original.is_empty():
                            self.log_message.emit(f"Dados vazios ou erro ao ler {current_item_description}. Pulando.", LogLevel.WARNING)
                            processed_items += 1
                            continue

                        # --- 1. Aplicar Mapeamento de Nomes e Filtro de Colunas (com Coalesce) ---
                        df_intermediate = df_original
                        if self.header_mapping:
                            # 1. Agrupar colunas de origem por seu nome final de destino
                            final_name_to_source = defaultdict(list)
                            for original_col_name in df_original.columns:
                                source_key = (original_col_name, file_path, sheet_name)
                                mapping_info = self.header_mapping.get(source_key)
                                if mapping_info and mapping_info.get("include", False):
                                    final_name = mapping_info.get("final_name")
                                    final_name_to_source[final_name].append(original_col_name)
                            # final_name_to_original_map = {}
                            # for original_name, mapping_info in self.header_mapping.items():
                            #     if mapping_info.get("include", False):
                            #         final_name = mapping_info.get("final_name", original_name)
                            #         if final_name not in final_name_to_original_map:
                            #             final_name_to_original_map[final_name] = []
                            #         final_name_to_original_map[final_name].append(original_name)

                            # 2. Construir as expressões de seleção usando coalesce quando necessário
                            select_expressions = []
                            df_original_columns = set(df_original.columns) # Para buscas mais rápidas

                            for final_name, original_cols_list in final_name_to_source.items():
                                # Filtrar a lista para incluir apenas as colunas que existem no DataFrame atual
                                # existing_original_cols = [col for col in original_cols_list if col in df_original_columns]
                                
                                # if not existing_original_cols:
                                #     continue # Nenhuma das colunas de origem para este nome final existe neste arquivo/aba

                                # Se apenas uma coluna de origem existe, faz um alias simples.
                                # Se mais de uma, usa coalesce para combinar os dados.
                                if not original_cols_list:
                                    continue
                                if len(original_cols_list) > 1:
                                    self.log_message.emit(f"Combinando colunas {original_cols_list} em '{final_name}' para {current_item_description}", LogLevel.INFO)
                                    expr = pl.coalesce(original_cols_list).alias(final_name)
                                else:
                                    self.log_message.emit(f"Combinando colunas {original_cols_list} em '{final_name}' para {current_item_description}", LogLevel.INFO)
                                    expr = pl.col(original_cols_list[0]).alias(final_name)
                                
                                select_expressions.append(expr)

                            # Se, após o mapeamento, não sobrar nenhuma expressão, pular o arquivo/aba
                            if not select_expressions:
                                self.log_message.emit(f"Nenhuma coluna do arquivo {current_item_description} corresponde ao mapeamento. Pulando.", LogLevel.WARNING)
                                processed_items += 1
                                continue

                            # 3. Executar a seleção no DataFrame
                            df_intermediate = df_original.select(select_expressions)
                        
                        if df_intermediate.width == 0:
                            self.log_message.emit(f"Nenhuma coluna restante em {current_item_description} após mapeamento de nomes. Pulando.", LogLevel.WARNING)
                            processed_items += 1; continue
                        
                        # --- 2. Aplicar Tipagem Especificada pelo Usuário ---
                        df_typed = df_intermediate
                        if self.header_mapping:
                            # Precisamos iterar sobre as colunas FINAIS do df_intermediate
                            # e encontrar a regra de tipagem correspondente no header_mapping
                            # (que é chaveado pelo nome ORIGINAL).
                            
                            # Criar um mapa reverso: {final_name: original_name_info_from_mapping}
                            # ou mais direto: {final_name: type_str_from_mapping}
                            final_name_to_type_str = {}
                            for original_h, map_details in self.header_mapping.items():
                                if map_details.get("include"):
                                    final_name_to_type_str[map_details.get("final_name", original_h)] = map_details.get("type_str")

                            casting_expressions = []
                            for final_col_name in df_typed.columns: # Iterar sobre colunas já mapeadas/renomeadas
                                type_str = final_name_to_type_str.get(final_col_name)
                                
                                if type_str and type_str != DATA_TYPES_OPTIONS[0]: # Se não for "Automático/String"
                                    polars_type = TYPE_STRING_TO_POLARS.get(type_str)
                                    if polars_type:
                                        self.log_message.emit(f"Convertendo coluna '{final_col_name}' para {type_str} em {current_item_description}", LogLevel.INFO)
                                        casting_expressions.append(pl.col(final_col_name).cast(polars_type, strict=False))
                                    else: # Tipo string não encontrado no mapa (não deve acontecer)
                                        casting_expressions.append(pl.col(final_col_name)) # Manter como está
                                else: # "Automático/String" ou tipo não mapeado
                                    casting_expressions.append(pl.col(final_col_name)) # Manter como está
                            
                            if casting_expressions: # Se houver alguma expressão (sempre haverá se df_typed não for vazio)
                                df_typed = df_typed.select(casting_expressions) # .select() é mais seguro que .with_columns() para recriar
                    
                        # --- 4. Aplicar Filtros (com lógica hierárquica E/OU) ---
                        df_filtered = df_typed
                        if self.filter_rules:
                            # from collections import defaultdict

                            # Definir quais operadores são para exclusão
                            EXCLUSION_OPERATORS = {"Diferente de", "Não contém"}

                            grouped_rules = defaultdict(list)
                            for rule in self.filter_rules:
                                if rule.get("column"):
                                    grouped_rules[rule["column"]].append(rule)

                            final_expressions_to_and = []
                            df_schema = df_filtered.schema

                            for col_name, rules_for_col in grouped_rules.items():
                                if col_name not in df_schema:
                                    continue

                                inclusion_exprs = []
                                exclusion_exprs = []
                                col_type = df_schema[col_name]

                                # 1. Separar regras em Inclusão e Exclusão
                                for rule in rules_for_col:
                                    operator = rule.get("operator")
                                    value = rule.get("value") # Pega o valor (pode ser string ou lista)

                                    target_list = exclusion_exprs if operator in EXCLUSION_OPERATORS else inclusion_exprs

                                    # Adicionado para pular regras incompletas
                                    if operator is None or value is None:
                                        continue
                                    
                                    try:
                                        polars_col = pl.col(col_name)
                                        expr = None
                                        
                                        if operator in OPERATORS_NO_VALUE:
                                            if operator == "Está em branco": expr = polars_col.is_null()
                                            elif operator == "Não está em branco": expr = polars_col.is_not_null()
                                        
                                        elif operator == "Entre":
                                            if isinstance(value, list) and len(value) == 2:
                                                min_val_str, max_val_str = value
                                                # Strip é aplicado aqui, onde sabemos que são strings
                                                if min_val_str.strip() and max_val_str.strip():
                                                    lit_min = pl.lit(min_val_str.strip()).cast(col_type, strict=False)
                                                    lit_max = pl.lit(max_val_str.strip()).cast(col_type, strict=False)
                                                    expr = polars_col.is_between(lit_min, lit_max)

                                        # Garante que o valor é uma string antes de usar o .strip()
                                        elif isinstance(value, str) and value.strip():
                                            value_str = value.strip()
                                            lit_val = pl.lit(value_str).cast(col_type, strict=False)

                                            if operator == "Igual a": expr = (polars_col == lit_val)
                                            elif operator == "Diferente de": expr = (polars_col != lit_val)
                                            elif operator == "Maior que": expr = (polars_col > lit_val)
                                            elif operator == "Menor que": expr = (polars_col < lit_val)
                                            elif col_type == pl.String:
                                                if operator == "Contém": expr = polars_col.str.contains(value_str, literal=True)
                                                elif operator == "Não contém": expr = ~polars_col.str.contains(value_str, literal=True)
                                                elif operator == "Começa com": expr = polars_col.str.starts_with(value_str)
                                                elif operator == "Termina com": expr = polars_col.str.ends_with(value_str)
                                        
                                        if expr is not None:
                                            target_list.append(expr)

                                    except Exception as e_filter:
                                        self.log_message.emit(f"Não foi possível aplicar a regra de filtro '{col_name} {operator} {value}': {e_filter}", LogLevel.WARNING)
                                
                                # 2. Construir a expressão final para esta coluna
                                col_final_expr = None
                                
                                # Combinar todas as expressões de inclusão com OU (OR)
                                final_inclusion_expr = pl.any_horizontal(inclusion_exprs) if len(inclusion_exprs) > 1 else (inclusion_exprs[0] if inclusion_exprs else None)
                                
                                # Combinar todas as expressões de exclusão com E (AND)
                                final_exclusion_expr = pl.all_horizontal(exclusion_exprs) if len(exclusion_exprs) > 1 else (exclusion_exprs[0] if exclusion_exprs else None)

                                # Juntar inclusão e exclusão com E (AND)
                                if final_inclusion_expr is not None and final_exclusion_expr is not None:
                                    col_final_expr = final_inclusion_expr & final_exclusion_expr
                                elif final_inclusion_expr is not None:
                                    col_final_expr = final_inclusion_expr
                                elif final_exclusion_expr is not None:
                                    col_final_expr = final_exclusion_expr

                                if col_final_expr is not None:
                                    final_expressions_to_and.append(col_final_expr)

                            # 3. Aplicar os filtros finais combinados com E (AND)
                            if final_expressions_to_and:
                                rows_before = df_filtered.height
                                df_filtered = df_filtered.filter(final_expressions_to_and)
                                rows_after = df_filtered.height
                                self.log_message.emit(f"Filtro aplicado em {current_item_description}. Linhas restantes: {rows_after} de {rows_before}.", LogLevel.INFO)
                        
                        # --- Fim do Bloco de Filtros ---

                        # --- 3. Adicionar Coluna de Origem --- 
                        file_name_only = os.path.basename(file_path)
                        source_name = f"{file_name_only} ({sheet_name})" if sheet_name else file_name_only
                        
                        df_final_for_list = df_filtered.with_columns( # <-- Usa df_filtered
                            pl.lit(source_name).alias("Origem")
                        )

                        all_dataframes_processed.append(df_final_for_list)

                    except Exception as e:
                        self.log_message.emit(f"Erro ao processar (ler/mapear/tipar) {current_item_description}: {e}", LogLevel.ERROR)
                    
                    processed_items += 1
                    progress = int((processed_items / total_items) * 100) if total_items > 0 else 0
                    self.progress_updated.emit(progress)

                if not self.is_running: break 

            if not self.is_running:
                 self.log_message.emit("Consolidação cancelada.", LogLevel.WARNING)
                 self.finished.emit(False, "Cancelado"); return

            if not all_dataframes_processed:
                self.log_message.emit("Nenhum dado após processamento.", LogLevel.WARNING)
                self.finished.emit(False, "Nenhum dado processado."); return

            # --- Harmonização de Tipos (Pós-Tipagem do Usuário e Mapeamento) ---
            self.log_message.emit("Harmonizando tipos (2ª passagem) entre arquivos processados...", LogLevel.INFO)
            
            # 1. Coletar todos os tipos para cada nome de coluna final único
            #    em todos os DataFrames processados.
            column_all_types_globally = {} # {final_col_name: set_of_dtypes}
            for df_processed in all_dataframes_processed:
                 for col_name, dtype in df_processed.schema.items():
                     if col_name not in column_all_types_globally:
                         column_all_types_globally[col_name] = set()
                     column_all_types_globally[col_name].add(dtype)
            
            # 2. Determinar o tipo alvo para cada coluna globalmente
            global_target_types = {} # {final_col_name: target_polars_type}
            for final_col_name, dtypes_set in column_all_types_globally.items():
                is_int_present = any(t.is_integer() for t in dtypes_set)
                is_float_present = any(t.is_float() for t in dtypes_set)
                is_string_present = any(t == pl.String or t == pl.Utf8 for t in dtypes_set)
                is_temporal_present = any(t.is_temporal() for t in dtypes_set)
                is_boolean_present = any(t == pl.Boolean for t in dtypes_set)
                is_null_present = any(t == pl.Null for t in dtypes_set) # Null type

                target_type_for_col = None

                # Regra de Prioridade para determinar o tipo alvo:
                if is_string_present: # Se String estiver presente, tudo vira String
                    target_type_for_col = pl.String
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global String (devido à presença de String).", LogLevel.INFO)
                elif is_temporal_present and (is_int_present or is_float_present or is_boolean_present): # Temporal com outros não-string -> String
                    target_type_for_col = pl.String
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global String (conflito Temporal com Numérico/Booleano).", LogLevel.INFO)
                elif is_boolean_present and (is_int_present or is_float_present): # Booleano com Numérico -> String
                    target_type_for_col = pl.String
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global String (conflito Booleano com Numérico).", LogLevel.INFO)
                elif is_float_present: # Se Float estiver presente (e não String), tudo vira Float
                    target_type_for_col = pl.Float64
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global Decimal (Float) (devido à presença de Float ou Int+Float).", LogLevel.INFO)
                elif is_int_present: # Se apenas Int (e talvez Null, Boolean que pode ser Int)
                    target_type_for_col = pl.Int64 # Se só há Int e Null, pode ser Int. Se Booleano, pode ser Int.
                    # Se Booleano estiver presente e quisermos ser mais específicos, poderíamos ter mais regras
                    # Mas Int64 pode acomodar Booleanos como 0/1 se o cast funcionar.
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global Inteiro.", LogLevel.INFO)
                elif is_temporal_present: # Apenas Temporal (e talvez Null)
                    # Se houver múltiplos tipos temporais (Date, Datetime, Duration), escolher o mais geral (Datetime?) ou String.
                    # Por simplicidade, se só Temporal, manter (o primeiro que encontrar, ou o mais comum).
                    # Para ser seguro, se houver vários tipos temporais, converter para String ou o tipo mais abrangente.
                    # Esta parte pode precisar de mais refinamento se você tiver mistura de Date, Datetime, etc.
                    # Vamos pegar o primeiro tipo temporal encontrado como exemplo, ou default para pl.Date.
                    first_temporal_type = next((t for t in dtypes_set if t.is_temporal()), pl.Date)
                    target_type_for_col = first_temporal_type
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global {first_temporal_type} (apenas Temporal).", LogLevel.INFO)
                elif is_boolean_present: # Apenas Booleano (e talvez Null)
                    target_type_for_col = pl.Boolean
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global Booleano.", LogLevel.INFO)
                # Se for apenas Null, ou tipos não cobertos, ele permanecerá None, e o cast não será aplicado abaixo
                # ou podemos definir um default como String.
                elif is_null_present and len(dtypes_set) == 1: # Apenas Null type
                     # Deixar como está por enquanto, o concat pode lidar com coluna toda Null.
                     # Ou, se quisermos ser proativos:
                     # target_type_for_col = pl.String
                     # self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global String (era apenas Null).", LogLevel.INFO)
                     pass # Não define target_type, o cast não será aplicado para esta coluna se ela for só Null

                if target_type_for_col:
                    global_target_types[final_col_name] = target_type_for_col
            
            # 3. Aplicar o tipo alvo global a cada DataFrame
            harmonized_dataframes_final_pass = [] 
            for df_to_harmonize in all_dataframes_processed:
                 df_modified_this_pass = df_to_harmonize
                 
                 expressions_to_apply = []
                 for col_name_in_df in df_to_harmonize.columns:
                     target_type = global_target_types.get(col_name_in_df)
                     current_type = df_to_harmonize.schema.get(col_name_in_df)

                     if target_type and current_type != target_type:
                         # Só aplicar cast se o tipo atual for diferente do alvo
                         expressions_to_apply.append(pl.col(col_name_in_df).cast(target_type, strict=False).alias(col_name_in_df))
                         self.log_message.emit(f"Aplicando tipo alvo '{target_type}' à coluna '{col_name_in_df}' (era '{current_type}').", LogLevel.INFO)
                     else:
                         # Manter a coluna como está (ou porque não há tipo alvo ou já é o tipo alvo)
                         expressions_to_apply.append(pl.col(col_name_in_df))
                 
                 if expressions_to_apply: # Se houver colunas no DF (sempre deve haver se chegou aqui)
                     df_modified_this_pass = df_to_harmonize.select(expressions_to_apply)
                 
                 harmonized_dataframes_final_pass.append(df_modified_this_pass)
            
            final_dataframes_to_concat = harmonized_dataframes_final_pass
            # --- Fim Harmonização (2ª passagem) ---
            # dataframes_final_pass_no_null_type = []
            # for df in dataframes_final_pass_no_null_type:
            #     expression_for_null_cast = []
            #     needs_null_type_casting = False
            #     for col_name in df.columns:
            #         if df[col_name].dtype == pl.Null:
            #             expression_for_null_cast.append(pl.col(col_name).cast(pl.String).alias(col_name))
            #             needs_null_type_casting = True
            #         else:
            #             expression_for_null_cast.append(pl.col(col_name))
            #     if needs_null_type_casting and expression_for_null_cast:
            #         dataframes_final_pass_no_null_type.append(df.select(expression_for_null_cast))
            #     else:
            #         final_dataframes_to_concat = dataframes_final_pass_no_null_type
            self.log_message.emit("Concatenando dados processados...", LogLevel.INFO)
            try:
                consolidated_df = pl.concat(final_dataframes_to_concat, how="diagonal")
                # --- Reordenar Coluna "Origem" para o Final ---
                if "Origem" in consolidated_df.columns:
                    # Pega todas as colunas, exceto "Origem"
                    all_other_columns = [col for col in consolidated_df.columns if col != "Origem"]
                    # Cria a nova ordem com "Origem" no final
                    new_column_order = all_other_columns + ["Origem"]
                    # Seleciona as colunas na nova ordem
                    consolidated_df = consolidated_df.select(new_column_order)
                
                # --- Remoção de duplicatas ---
                removed_duplicates_df = None
                key_columns = self.duplicates_config.get("key_columns", [])
                generate_report = self.duplicates_config.get("generate_report", False)
                if key_columns:
                    rows_before = consolidated_df.height
                    self.log_message.emit(f"Removendo duplicatas com base nas chaves: {', '.join(key_columns)}...", LogLevel.INFO)
                    df_with_index = consolidated_df.with_row_index("__temp_index__")
                    unique_rows = df_with_index.unique(subset = key_columns, keep = 'first')
                    if generate_report:
                        removed_duplicates_df = df_with_index.join(unique_rows, on = "__temp_index__", how = "anti").drop("__temp_index__")
                    consolidated_df = unique_rows.drop("__temp_index__")
                    rows_after = consolidated_df.height
                    self.log_message.emit(f"{rows_before - rows_after} linhas duplicadas foram removidas. Linhas restantes: {rows_after}", LogLevel.SUCCESS)
                    if removed_duplicates_df is not None and not removed_duplicates_df.is_empty():
                        self.log_message.emit(f"Uma aba com as {removed_duplicates_df.height} linhas removidas será gerada.", LogLevel.INFO)
                
                # --- Aplicar Regras da Tabela de Resumo (Pivot) ---
                pivot_df = None # DataFrame para a tabela de resumo
                if self.pivot_rules and self.pivot_rules.get("group_by") and self.pivot_rules.get("aggregations"):
                    self.log_message.emit("Criando Tabela de Resumo (em memória)...", LogLevel.INFO)
                    try:
                        group_by_cols = self.pivot_rules['group_by']
                        aggregations = self.pivot_rules['aggregations']

                        op_map = {
                            "Soma": pl.sum, "Média": pl.mean, "Contagem": pl.count,
                            "Mínimo": pl.min, "Máximo": pl.max,
                            "Contagem Única": lambda col: pl.col(col).n_unique()
                        }

                        agg_expressions = []
                        for rule in aggregations:
                            col_name = rule['column']
                            op_str = rule['operation']
                            
                            if col_name not in consolidated_df.columns:
                                self.log_message.emit(f"Coluna '{col_name}' da regra de resumo não encontrada. Pulando.", LogLevel.WARNING)
                                continue
                            
                            if op_str in op_map:
                                polars_func = op_map[op_str]
                                new_col_name = f"{col_name}_{op_str.replace(' ', '_')}"
                                agg_expressions.append(polars_func(col_name).alias(new_col_name))
                        
                        if agg_expressions:
                            pivot_df = consolidated_df.group_by(group_by_cols).agg(agg_expressions).sort(group_by_cols)
                            self.log_message.emit("Tabela de resumo criada com sucesso.", LogLevel.SUCCESS)

                    except Exception as e_pivot:
                        self.log_message.emit(f"Erro ao criar tabela de resumo: {e_pivot}. O resultado do resumo não será salvo.", LogLevel.ERROR)
                        pivot_df = None
                # --- FIM DO BLOCO DE PIVOT --
            except Exception as e: 
                 self.log_message.emit(f"Erro concatenação final: {e}", LogLevel.ERROR)
                 self.finished.emit(False, f"Erro concatenação: {e}"); return
            if self.output_format == "XLSX":
                illegal_xml_chars_re = r"[\u0000-\u0008\u000B\u000C\u000E-\u001F]"
                # Sanitiza ambos os dataframes
                consolidated_df = consolidated_df.with_columns(
                    pl.col(pl.String).str.replace_all(illegal_xml_chars_re, "")
                )
                if pivot_df is not None:
                    pivot_df = pivot_df.with_columns(
                        pl.col(pl.String).str.replace_all(illegal_xml_chars_re, "")
                    )
            
            self.log_message.emit(f"Salvando: {self.output_path}", LogLevel.INFO)
            only_pivot = self.pivot_rules.get("only_pivot", False)

            if self.output_format == "XLSX":
                try:
                    import xlsxwriter
                    workbook = xlsxwriter.Workbook(self.output_path, {'use_zip64': True})
                    
                    # Formatos
                    header_format = workbook.add_format({'font_name': 'Aptos', 'bold': True, 'font_color': 'white', 'bg_color': '#000000', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
                    data_format = workbook.add_format({'font_name': 'Aptos'})
                    group_by_header_format = workbook.add_format({'font_name': 'Aptos', 'bold': True, 'font_color': 'white', 'bg_color': '#000000', 'border': 1, 'align': 'center', 'valign': 'vcenter'})

                    total_rows_to_write = consolidated_df.height + (pivot_df.height if pivot_df is not None else 0)
                    total_rows_written = 0

                    # 1. Escrever a Tabela de Resumo (pivot_df), se existir
                    if pivot_df is not None:
                        self.log_message.emit("Escrevendo aba 'Tabela_Resumo'...", LogLevel.INFO)
                        worksheet_pivot = workbook.add_worksheet("Tabela_Resumo")
                        worksheet_pivot.freeze_panes('A2')
                        worksheet_pivot.set_zoom(70)
                        worksheet_pivot.hide_gridlines(2)
                        
                        group_by_cols = self.pivot_rules.get('group_by', [])
                        for col_idx, col_name in enumerate(pivot_df.columns):
                            fmt = group_by_header_format if col_name in group_by_cols else header_format
                            worksheet_pivot.write(0, col_idx, col_name, fmt)
                        
                        max_lengths_pivot = {col: max(len(str(col)), pivot_df.select(pl.col(col).cast(pl.String).str.len_chars().max()).item() or 1) for col in pivot_df.columns}
                        for col_idx, col_name in enumerate(pivot_df.columns):
                            width = min(max_lengths_pivot.get(col_name, len(col_name)) + 2, 60)
                            worksheet_pivot.set_column(col_idx, col_idx, width, data_format)
                        
                        for r_idx, row_tuple in enumerate(pivot_df.iter_rows(), start=1):
                            worksheet_pivot.write_row(r_idx, 0, row_tuple)
                            total_rows_written += 1
                    if not only_pivot:
                        # 2. Escrever os Dados Consolidados 
                        self.log_message.emit("Escrevendo aba(s) de 'Dados_Consolidados'...", LogLevel.INFO)
                        max_rows_per_sheet = 1_048_570
                        dfs_to_write = []
                        if consolidated_df.height > max_rows_per_sheet:
                            num_chunks = (consolidated_df.height + max_rows_per_sheet - 1) // max_rows_per_sheet
                            for i in range(num_chunks):
                                df_chunk = consolidated_df.slice(i * max_rows_per_sheet, max_rows_per_sheet)
                                dfs_to_write.append((f"Dados_Consolidados_{i+1}", df_chunk))
                        else:
                            dfs_to_write.append(("Dados_Consolidados", consolidated_df))

                        max_lengths_consolidated = {col: max(len(str(col)), consolidated_df.select(pl.col(col).cast(pl.String).str.len_chars().max()).item() or 1) for col in consolidated_df.columns}
                        if removed_duplicates_df is not None and not removed_duplicates_df.is_empty():
                            self.log_message.emit("Escrevendo aba 'Duplicatas_Removidas'...", LogLevel.INFO)
                            duplicates_header_format = workbook.add_format({'font_name': 'Aptos', 'bold': True, 'font_color': 'white', 'bg_color': '#C00000', 'border': 1, 'align': 'center', 'valign': 'vcenter'}) # Cabeçalho vermelho

                            # A lógica de chunking é importante aqui também
                            max_rows_per_sheet = 1_048_570
                            duplicates_to_write = []
                            if removed_duplicates_df.height > max_rows_per_sheet:
                                num_chunks = (removed_duplicates_df.height + max_rows_per_sheet - 1) // max_rows_per_sheet
                                for i in range(num_chunks):
                                    duplicates_to_write.append((f"Duplicatas_Removidas_{i+1}", removed_duplicates_df.slice(i * max_rows_per_sheet, max_rows_per_sheet)))
                            else:
                                duplicates_to_write.append(("Duplicatas_Removidas", removed_duplicates_df))
                            
                            max_lengths_duplicates = {col: max(len(str(col)), removed_duplicates_df.select(pl.col(col).cast(pl.String).str.len_chars().max()).item() or 1) for col in removed_duplicates_df.columns}

                            for sheet_name, df_chunk in duplicates_to_write:
                                worksheet = workbook.add_worksheet(sheet_name)
                                worksheet.freeze_panes('A2')
                                worksheet.set_zoom(70)
                                worksheet.hide_gridlines(2)
                                worksheet.write_row('A1', df_chunk.columns, duplicates_header_format)
                                worksheet.autofilter(0, 0, df_chunk.height, df_chunk.width - 1)
                                for col_idx, col_name in enumerate(df_chunk.columns):
                                    worksheet.set_column(col_idx, col_idx, min(max_lengths_duplicates.get(col_name, len(col_name)) + 2, 60), data_format)
                                for r_idx, row_tuple in enumerate(df_chunk.iter_rows(), start=1):
                                    worksheet.write_row(r_idx, 0, row_tuple)
                                    total_rows_written += 1
                                    if total_rows_written % 5000 == 0:
                                        self.progress_text_updated.emit(f"Escrevendo linha {total_rows_written:,} de {total_rows_to_write:,}...")
                                
                            
                            # max_lengths_duplicates = {col: max(len(str(col)), removed_duplicates_df.select(pl.col(col).cast(pl.String).str.len_chars().max()).item() or 1) for col in removed_duplicates_df.columns}
                        for sheet_name, df_chunk in dfs_to_write:
                            worksheet = workbook.add_worksheet(sheet_name)
                            worksheet.freeze_panes('A2')
                            worksheet.set_zoom(70)
                            worksheet.hide_gridlines(2)
                            worksheet.write_row('A1', df_chunk.columns, header_format)
                            worksheet.autofilter(0, 0, df_chunk.height, df_chunk.width - 1)
                            for col_idx, col_name in enumerate(df_chunk.columns):
                                width = min(max_lengths_consolidated.get(col_name, len(col_name)) + 2, 60)
                                worksheet.set_column(col_idx, col_idx, width, data_format)
                            
                            for r_idx, row_tuple in enumerate(df_chunk.iter_rows(), start=1):
                                worksheet.write_row(r_idx, 0, row_tuple)
                                total_rows_written += 1
                                if total_rows_written % 5000 == 0:
                                    self.progress_text_updated.emit(f"Escrevendo linha {total_rows_written:,} de {total_rows_to_write:,}...")
                        
                        self.progress_text_updated.emit(f"Finalizando escrita de {total_rows_to_write:,} linhas...")
                    workbook.close()

                except Exception as e_save_excel:
                    self.log_message.emit(f"Erro ao salvar arquivo Excel com XlsxWriter: {e_save_excel}", LogLevel.ERROR)
                    self.finished.emit(False, f"Erro ao salvar Excel: {e_save_excel}")
                    return

            elif self.output_format in ["CSV", "Parquet"]:
                df_to_save = pivot_df if pivot_df is not None else consolidated_df
                if pivot_df is not None:
                    self.log_message.emit(f"Salvando resultado da Tabela de Resumo em {self.output_format}.", LogLevel.INFO)
                
                if self.output_format == "CSV":
                    df_to_save.write_csv(self.output_path, separator='|')
                elif self.output_format == "Parquet":
                    df_to_save.write_parquet(self.output_path, compression='zstd')

            self.progress_updated.emit(100)
            self.log_message.emit(f"Concluído! Salvo em: {self.output_path}", LogLevel.SUCCESS)
            self.finished.emit(True, f"Salvo em: {self.output_path}")

        except Exception as e:
            self.log_message.emit(f"Erro inesperado consolidação: {e}", LogLevel.ERROR)
            self.finished.emit(False, f"Erro: {e}")

    def stop(self): # stop() permanece o mesmo
        self.is_running = False
        self.log_message.emit("Tentativa de parada da consolidação solicitada...", LogLevel.INFO)

class SheetLoadingWorker(QThread):
    finished = Signal(str, list, str)  # file_path, sheet_names_list, error_message_or_None

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path
        self.is_running = True

    def run(self):
        if not self.is_running:
            self.finished.emit(self.file_path, [], "Cancelado antes de iniciar.")
            return

        sheet_names_from_file = []
        error_message = None
        try:
            if not self.is_running: # Checar novamente
                raise InterruptedError("Carregamento de abas cancelado.")

            if self.file_path.lower().endswith(".xlsx"):
                # Usar read_only=True para performance, data_only=True para evitar carregar fórmulas complexas se não necessário
                workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
                sheet_names_from_file = workbook.sheetnames
                workbook.close() # É bom fechar o workbook
            elif self.file_path.lower().endswith(".xls"):
                workbook = xlrd.open_workbook(self.file_path, on_demand=True)
                sheet_names_from_file = workbook.sheet_names()
            # Não precisa de else, pois a thread só é chamada para .xlsx/.xls
        except InterruptedError as ie:
            error_message = str(ie)
        except Exception as e:
            error_message = f"Erro ao ler abas do arquivo {os.path.basename(self.file_path)}: {e}"
        
        if self.is_running: # Só emite se não foi cancelado durante a operação
            self.finished.emit(self.file_path, sheet_names_from_file, error_message)

    def stop(self):
        self.is_running = False

class SheetAnalysisWorker(QThread):
    """
    Worker para analisar todos os arquivos Excel em uma lista e retornar
    um cache com todas as abas por arquivo e um conjunto de nomes de abas únicos.
    """
    # Sinal emite: {cache_de_abas}, {abas_unicas}, mensagem_de_erro
    finished = Signal(dict, set, str)

    def __init__(self, excel_files_paths):
        super().__init__()
        self.excel_files_paths = excel_files_paths
        self.is_running = True

    def run(self):
        all_sheets_cache = {} # Ex: {'caminho/arquivo1.xlsx': ['Aba1', 'Aba2'], ...}
        unique_sheet_names = set()
        error_message = None

        try:
            for file_path in self.excel_files_paths:
                if not self.is_running:
                    raise InterruptedError("Análise de abas cancelada.")
                
                try:
                    sheet_names = []
                    # Usamos openpyxl diretamente pois é otimizado para apenas ler nomes de abas
                    if file_path.lower().endswith(".xlsx"):
                        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        sheet_names = workbook.sheetnames
                        workbook.close()
                    elif file_path.lower().endswith(".xls"):
                        workbook = xlrd.open_workbook(file_path, on_demand=True)
                        sheet_names = workbook.sheet_names()
                    
                    if sheet_names:
                        all_sheets_cache[file_path] = sheet_names
                        unique_sheet_names.update(sheet_names)

                except Exception as e:
                    # Loga um erro para um arquivo específico mas continua o processo
                    print(f"AVISO: Não foi possível ler as abas de '{os.path.basename(file_path)}'. Erro: {e}")
            
            if self.is_running:
                self.finished.emit(all_sheets_cache, unique_sheet_names, None)

        except InterruptedError as ie:
            error_message = str(ie)
        except Exception as e:
            error_message = f"Um erro inesperado ocorreu durante a análise de abas: {e}"
        
        if error_message:
            self.finished.emit({}, set(), error_message)

    def stop(self):
        self.is_running = False

class HeaderAnalysisWorker(QThread):
    '''Worker para os cabeçalhos'''
    finished = Signal(list, object)
    progress_log = Signal(str, LogLevel)

    def __init__(self, files_and_sheets_config, delimiter):
        super().__init__()
        self.files_and_sheets_config = files_and_sheets_config
        self.delimiter = delimiter
        self.is_running = True

    def _get_series_profile(self, series: pl.Series):
        non_null_series = series.filter(series.is_not_null() & (series.str.strip_chars() != ""))
        if non_null_series.is_empty():
            return {"dtype": pl.String, "null_ratio": series.is_null().mean()}
        try:
            if non_null_series.cast(pl.Int64, strict=True).is_not_null().all():
                return {"dtype": pl.Int64, "null_ratio": series.is_null().mean()}
        except (Exception, pl.exceptions.PanicException): pass
        try:
            if non_null_series.cast(pl.Float64, strict=True).is_not_null().all():
                return {"dtype": pl.Float64, "null_ratio": series.is_null().mean()}
        except (Exception, pl.exceptions.PanicException): pass
        try:
            if non_null_series.str.to_datetime(strict=True, exact=False, cache=False).is_not_null().all():
                return {"dtype": pl.Datetime, "null_ratio": series.is_null().mean()}
        except (Exception, pl.exceptions.PanicException): pass
        return {"dtype": pl.String, "null_ratio": series.is_null().mean()}

    def run(self):
        if not self.is_running:
            self.finished.emit([], InterruptedError("Análise cancelada."))
            return

        all_column_fingerprints = []
        n_sample_rows, n_preread_rows = 200, 20

        try:
            for file_path, selected_sheets in self.files_and_sheets_config:
                if not self.is_running: raise InterruptedError("Análise cancelada.")

                self.progress_log.emit(f"Analisando: {os.path.basename(file_path)}...", LogLevel.INFO)
                
                sheets_to_iterate = selected_sheets if selected_sheets is not None else [None]
                for sheet_name in sheets_to_iterate:
                    if not self.is_running: raise InterruptedError("Análise cancelada.")
                    try:
                        pre_read_df = None
                        if file_path.lower().endswith((".csv", ".txt")):
                            pre_read_df = pl.read_csv(source=file_path, has_header=False, n_rows=n_preread_rows, separator=self.delimiter, encoding='latin-1', ignore_errors=True, infer_schema = False, truncate_ragged_lines = True, quote_char = None)
                        elif file_path.lower().endswith((".xlsx", ".xls")):
                            pre_read_df = pl.read_excel(source=file_path, sheet_name=sheet_name, has_header = False, infer_schema_length = 0).head(n_preread_rows)
                        
                        if pre_read_df is None or pre_read_df.is_empty(): continue
                        
                        header_row_index = _find_header_row_index(pre_read_df, n_preread_rows)
                        header_names_raw = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(pre_read_df.row(header_row_index))]
                        header_names = _make_headers_unique(header_names_raw)
                        data_rows_df = pre_read_df.slice(offset=header_row_index + 1).head(n_sample_rows)
                        
                        if data_rows_df.is_empty(): continue
                        
                        rename_mapping = {old_name: new_name for old_name, new_name in zip(data_rows_df.columns, header_names)}
                        sample_df = data_rows_df.rename(rename_mapping)

                        for col_name in sample_df.columns:
                            try: # <-- INÍCIO DO BLOCO DE BLINDAGEM
                                series = sample_df[col_name]
                                profile = self._get_series_profile(series)
                                fingerprint = {
                                    "source_tuple": (col_name, file_path, sheet_name),
                                    "normalized_name": _normalize_header_name(col_name),
                                    "dtype": profile["dtype"],
                                    "null_ratio": profile["null_ratio"],
                                }
                            except Exception as e_profile: # <-- CAPTURA O "PANIC"
                                # Se a análise da coluna falhar, cria um perfil "seguro"
                                self.progress_log.emit(f"Falha ao analisar coluna '{col_name}' em '{os.path.basename(file_path)}'. Tratando como texto. Erro: {e_profile}", LogLevel.WARNING)
                                fingerprint = {
                                    "source_tuple": (col_name, file_path, sheet_name),
                                    "normalized_name": _normalize_header_name(col_name),
                                    "dtype": pl.String, # Tipo de dado seguro
                                    "null_ratio": 0.0,
                                }
                            all_column_fingerprints.append(fingerprint)
                    except Exception:
                        continue

            # --- ALGORITMO DE AGRUPAMENTO DEFINITIVO ---
            groups_by_name = defaultdict(list)
            for fp in all_column_fingerprints:
                groups_by_name[fp["normalized_name"]].append(fp)
            
            final_groups = []
            for normalized_name, fp_list in groups_by_name.items():
                if len(fp_list) == 1:
                    final_groups.append([fp_list[0]["source_tuple"]])
                    continue

                sub_groups_by_type = defaultdict(list)
                for fp in fp_list:
                    type_key = "numeric" if fp["dtype"].is_numeric() else str(fp["dtype"])
                    sub_groups_by_type[type_key].append(fp["source_tuple"])
                
                for type_key, source_tuples in sub_groups_by_type.items():
                    final_groups.append(source_tuples)

            if self.is_running:
                self.finished.emit(final_groups, None)
        except Exception as e:
            self.finished.emit([], e)

    def stop(self):
        self.is_running = False

# Exceção customizada para interrupção
class InterruptedError(Exception):
    pass

class PolarsTableModel(QAbstractTableModel):
    def __init__(self, data=None):
        super().__init__()
        self._data = data if data is not None else pl.DataFrame()

    def rowCount(self, parent=None):
        return self._data.height

    def columnCount(self, parent=None):
        return self._data.width

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.DisplayRole:
            try:
                # Polars pode retornar um único valor ou uma Series.
                # Se for Series (ao pegar uma célula), pegue o primeiro valor.
                value = self._data[index.row(), index.column()]
                if isinstance(value, pl.Series): # Acesso a célula pode retornar Series de 1 elemento
                    return str(value[0]) if value.len() > 0 else ""
                return str(value) # Converte para string para exibição
            except Exception:
                return "" # Em caso de erro ao acessar, retorna string vazia
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal: # Cabeçalhos das colunas
                return str(self._data.columns[section])
            if orientation == Qt.Vertical: # Cabeçalhos das linhas (números)
                return str(section + 1)
        return None

    def load_data(self, new_data: pl.DataFrame):
        self.beginResetModel()
        self._data = new_data if new_data is not None else pl.DataFrame()
        self.endResetModel()

    def clear_data(self):
        self.load_data(pl.DataFrame())

class SplitGroupDialog(QDialog):
    """Um diálogo para dividir um grupo de cabeçalhos."""
    def __init__(self, group_to_split, parent = None):
        super().__init__(parent)
        self.setWindowTitle("Dividir Grupo de Cabeçalhos")
        self.setMinimumWidth(500)
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(QLabel("Marque os cabeçalhos que você deseja mover para um novo grupo:"))
        self.list_widget = QListWidget()
        for original_name, file_path, sheet_name in group_to_split:
            # Formata o texto do item para ser informativo
            source_text = os.path.basename(file_path)
            if sheet_name:
                source_text += f" ({sheet_name})"
            item_text = f"'{original_name}' (Fonte: {source_text})"
            list_item = QListWidgetItem(item_text)
            list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
            list_item.setCheckState(Qt.Unchecked)
            # Armazena a tupla original no item para referência posterior
            list_item.setData(Qt.UserRole, (original_name, file_path, sheet_name))
            self.list_widget.addItem(list_item)
        self.layout.addWidget(self.list_widget)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        self.layout.addWidget(button_box)
    
    def get_selected_to_split(self):
        """Retorna uma lista das tuplas de origem que foram marcadas
        para serem divididas."""
        to_split = []
        # all_items = self.list_widget.findItems("*", Qt.MatchWildcard)
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            is_checked = (item.checkState() == Qt.Checked)
            # print(f"Item {i} ('{item.text()}'): Marcado? {is_checked}")
            # if item.checkState() == Qt.Checked:
            if is_checked:
                to_split.append(item.data(Qt.UserRole))
            # print(f"--- DEBUG: Itens que serão retornados para divisão: {to_split} ---\n")
        return to_split

class HeaderMappingDialog(QDialog):
    def __init__(self, suggested_groups, parent=None, existing_mapping=None, existing_duplicate_keys=None):
        super().__init__(parent)
        self.setWindowTitle("Mapeamento e Agrupamento de Cabeçalhos")
        self.setMinimumSize(950, 600)

        self.groups = suggested_groups
        # O existing_mapping pode ser usado no futuro para pré-preencher, por enquanto simplificamos
        
        layout = QVBoxLayout(self)
        description_label = QLabel(
            "O algoritmo sugeriu os grupos abaixo. Confirme, divida ou mescle os grupos e defina o mapeamento final."
        )
        layout.addWidget(description_label)

        # --- Layout de Pesquisa ---
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Pesquisar Grupo:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Digite para filtrar por nome do grupo ou membros...")
        self.search_input.textChanged.connect(self.filter_table)
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)

        # --- Layout para botões de ação (com os novos botões) ---
        actions_layout = QHBoxLayout()
        self.split_group_button = QPushButton("Dividir Grupo...")
        self.split_group_button.clicked.connect(self.split_selected_group)
        self.merge_groups_button = QPushButton("Mesclar Grupos")
        self.merge_groups_button.clicked.connect(self.merge_selected_groups)
        self.batch_type_change_button = QPushButton("Alterar Tipo para Selecionados")
        self.batch_type_change_button.clicked.connect(self.change_type_for_selected)
        
        self.mark_all_button = QPushButton("Marcar Visíveis")
        self.mark_all_button.clicked.connect(lambda: self.mark_or_unmark_all_visible(check=True))
        self.unmark_all_button = QPushButton("Desmarcar Visíveis")
        self.unmark_all_button.clicked.connect(lambda: self.mark_or_unmark_all_visible(check=False))

        actions_layout.addWidget(self.split_group_button)
        actions_layout.addWidget(self.merge_groups_button)
        actions_layout.addWidget(self.batch_type_change_button)
        actions_layout.addStretch() # Espaçador
        actions_layout.addWidget(self.mark_all_button)
        actions_layout.addWidget(self.unmark_all_button)
        layout.addLayout(actions_layout)

        # --- Tabela Principal ---
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(4)
        self.table_widget.setHorizontalHeaderLabels(["Grupo Sugerido / Cabeçalho", "Cabeçalho Final", 'Tipo de Dados', "Incluir?"])
        self.table_widget.setSelectionMode(QAbstractItemView.ExtendedSelection) # Permite selecionar múltiplas linhas
        
        self.populate_table() # Popula a tabela com os grupos iniciais

        self.table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table_widget.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table_widget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        layout.addWidget(self.table_widget)

        # --- Seção para Remoção de Duplicatas ---
        duplicates_box = QGroupBox("Remoção de Duplicatas (Opcional)")
        duplicates_layout = QVBoxLayout(duplicates_box)
        duplicates_layout.addWidget(QLabel("Selecione as colunas-chave para remover linhas duplicadas. Uma linha será removida se a combinação de valores nestas colunas já existir."))
        
        self.report_duplicates_checkbox = QCheckBox("Gerar aba com as linhas duplicadas que foram removidas.")
        self.report_duplicates_checkbox.setChecked(True)
        duplicates_layout.addWidget(self.report_duplicates_checkbox)

        self.duplicate_check_list = QListWidget()
        self.duplicate_check_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        # Popula a lista com os nomes finais sugeridos
        final_header_names = [group[0][0] for group in self.groups]
        self.duplicate_check_list.addItems(sorted(final_header_names))
        
        # Pré-seleciona chaves existentes, se houver
        if existing_duplicate_keys:
            for key_col in existing_duplicate_keys:
                items = self.duplicate_check_list.findItems(key_col, Qt.MatchExactly)
                if items:
                    items[0].setSelected(True)

        duplicates_layout.addWidget(self.duplicate_check_list)
        layout.addWidget(duplicates_box)

        # Botões OK e Cancelar
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

    def get_duplicate_check_columns(self):
        """Retorna a lista de colunas selecionadas para a verificação de duplicatas."""
        return [item.text() for item in self.duplicate_check_list.selectedItems()]
    
    def get_duplicates_config(self):
        """
        Retorna a configuração completa para remoção de duplicatas.
        """
        key_columns = [item.text() for item in self.duplicate_check_list.selectedItems()]
        generate_report = self.report_duplicates_checkbox.isChecked()
        return {
            "key_columns": key_columns,
            "generate_report": generate_report
        }

    def populate_table(self):
        """Limpa e preenche a tabela com base na lista self.groups atual."""
        self.table_widget.setRowCount(0) # Limpa a tabela
        self.table_widget.setRowCount(len(self.groups))

        for row, group_list in enumerate(self.groups):
            # group_list é uma lista de tuplas de origem, ex: [('CNPJ', 'f1', 's1'), ...]
            
            # Coluna 0: Representação do Grupo
            # Usa o primeiro nome original do grupo como nome de exibição
            display_name = group_list[0][0]
            item_group = QTableWidgetItem(display_name)
            item_group.setFlags(item_group.flags() & ~Qt.ItemIsEditable)
            
            # Cria o texto do tooltip para mostrar todos os membros do grupo
            tooltip_text = "Membros do grupo:\n" + "\n".join(
                f"- '{orig_name}' (em {os.path.basename(path)}{f' | {s_name}' if s_name else ''})"
                for orig_name, path, s_name in group_list
            )
            item_group.setToolTip(tooltip_text)
            item_group.setData(Qt.UserRole, group_list) # Armazena os dados do grupo no item
            self.table_widget.setItem(row, 0, item_group)

            # Coluna 1: Cabeçalho Final (QLineEdit)
            final_name_edit = QLineEdit(display_name) # Sugere o nome do grupo como nome final
            self.table_widget.setCellWidget(row, 1, final_name_edit)

            # Coluna 2: Tipo de Dados (QComboBox)
            type_combo = QComboBox()
            type_combo.addItems(DATA_TYPES_OPTIONS)
            self.table_widget.setCellWidget(row, 2, type_combo)

            # Coluna 3: Incluir? (QCheckBox)
            checkbox, checkbox_widget = self._create_checkbox()
            checkbox.setChecked(True) # Incluir por padrão
            self.table_widget.setCellWidget(row, 3, checkbox_widget)

    def _create_checkbox(self):
        """Helper para criar um QCheckBox centralizado dentro de um widget."""
        checkbox_widget = QWidget()
        checkbox_layout = QHBoxLayout(checkbox_widget)
        checkbox = QCheckBox()
        checkbox_layout.addWidget(checkbox)
        checkbox_layout.setAlignment(Qt.AlignCenter)
        checkbox_layout.setContentsMargins(0, 0, 0, 0)
        return checkbox, checkbox_widget

    def filter_table(self, search_text):
        """Filtra as linhas da tabela com base no texto de pesquisa."""
        search_text_lower = search_text.lower().strip()

        for row in range(self.table_widget.rowCount()):
            group_item = self.table_widget.item(row, 0)
            if group_item:
                # Pesquisa no nome do grupo e no tooltip (que contém todos os membros)
                display_name = group_item.text().lower()
                tooltip_text = group_item.toolTip().lower()

                # A linha é visível se o texto de pesquisa estiver no nome ou no tooltip
                is_visible = (search_text_lower in display_name) or \
                             (search_text_lower in tooltip_text)

                self.table_widget.setRowHidden(row, not is_visible)

    def mark_or_unmark_all_visible(self, check=True):
        """Marca ou desmarca a caixa 'Incluir?' para todas as linhas VISÍVEIS na tabela."""
        for row in range(self.table_widget.rowCount()):
            # Aplica a alteração apenas se a linha não estiver oculta pela pesquisa
            if not self.table_widget.isRowHidden(row):
                checkbox_widget = self.table_widget.cellWidget(row, 3)
                if checkbox_widget:
                    checkbox = checkbox_widget.layout().itemAt(0).widget()
                    checkbox.setChecked(check)

    def change_type_for_selected(self):
        """Altera o tipo de dado para todas as linhas selecionadas na tabela."""
        selected_rows = sorted(list(set(index.row() for index in self.table_widget.selectedIndexes())))

        if not selected_rows:
            # Você pode adicionar um QMessageBox aqui para informar o usuário, se desejar
            return

        # Usa QInputDialog para pegar o novo tipo do usuário
        new_type, ok = QInputDialog.getItem(
            self,
            "Alterar Tipo de Dados em Lote",
            "Selecione o novo tipo de dados para os grupos selecionados:",
            DATA_TYPES_OPTIONS,  # A constante global com os tipos
            0,                   # Índice inicial
            False                # Não editável
        )

        if ok and new_type:
            for row in selected_rows:
                type_combo = self.table_widget.cellWidget(row, 2)
                if isinstance(type_combo, QComboBox):
                    type_combo.setCurrentText(new_type)

    def split_selected_group(self):
        """Abre o diálogo para dividir o grupo atualmente selecionado."""
        selected_rows = sorted(list(set(index.row() for index in self.table_widget.selectedIndexes())))
        if len(selected_rows) != 1:
            # Informar o usuário que apenas um grupo pode ser dividido por vez
            return

        row_to_split = selected_rows[0]
        group_item = self.table_widget.item(row_to_split, 0)
        group_data = group_item.data(Qt.UserRole)
        # print(f"DEBUG - Grupo para Dividir: {group_data}")

        if len(group_data) < 2: # Não pode dividir um grupo de 1
            return

        split_dialog = SplitGroupDialog(group_data, self)
        if split_dialog.exec() == QDialog.Accepted:
            items_to_move = split_dialog.get_selected_to_split()
            if not items_to_move: return

            # Atualizar self.groups: remover os itens do grupo antigo e criar um novo
            new_group = items_to_move
            self.groups.append(new_group)
            
            # Lista atualizada do grupo antigo
            updated_old_group = [item for item in group_data if item not in items_to_move]
            
            if not updated_old_group: # Se todos os itens foram movidos
                self.groups.pop(self.groups.index(group_data))
            else:
                self.groups[self.groups.index(group_data)] = updated_old_group
            
            self.populate_table() # Redesenha a tabela com os grupos atualizados

    def merge_selected_groups(self):
        """Mescla duas ou mais linhas (grupos) selecionadas em uma só."""
        selected_rows = sorted(list(set(index.row() for index in self.table_widget.selectedIndexes())), reverse=True)
        if len(selected_rows) < 2:
            # Informar o usuário que precisa selecionar pelo menos 2 grupos para mesclar
            return

        new_merged_group = []
        groups_to_remove = []
        for row in selected_rows:
            group_data = self.table_widget.item(row, 0).data(Qt.UserRole)
            new_merged_group.extend(group_data)
            groups_to_remove.append(group_data)
        
        # Remover os grupos antigos e adicionar o novo grupo mesclado
        for group in groups_to_remove:
            self.groups.remove(group)
        self.groups.insert(0, new_merged_group) # Adiciona no topo para fácil visualização

        self.populate_table() # Redesenha a tabela

    def get_mapping(self):
        """Lê a tabela e retorna o dicionário de mapeamento final."""
        final_mapping = {}
        for row in range(self.table_widget.rowCount()):
            # Obter os dados do grupo armazenados no item da Coluna 0
            group_data = self.table_widget.item(row, 0).data(Qt.UserRole)

            if not group_data:
                continue
            
            # Obter as configurações definidas pelo usuário para este grupo
            final_name = self.table_widget.cellWidget(row, 1).text().strip()
            type_str = self.table_widget.cellWidget(row, 2).currentText()
            include = self.table_widget.cellWidget(row, 3).layout().itemAt(0).widget().isChecked()

            if not final_name: # Fallback se o nome final for deixado em branco
                final_name = self.table_widget.item(row, 0).text()

            # "Desdobra" o grupo, aplicando a mesma regra a todos os seus membros
            # for original_name, _, _ in group_data:
            for source_tuple in group_data:
                final_mapping[source_tuple] = {
                    "final_name": final_name,
                    "type_str": type_str,
                    "include": include
                }
        return final_mapping

class HelpDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Guia do Usuário - DataFlow")
        self.setMinimumSize(800, 600)

        # --- Estrutura do Conteúdo da Ajuda ---
        # Chave: Título do Tópico, Valor: Conteúdo em HTML
        self.help_content = {
            "Visão Geral": """
                <h1>Bem-vindo ao DataFlow!</h1>
                <p>Esta ferramenta foi projetada para automatizar o processo de unificar múltiplos arquivos
                (Excel, CSV, TXT) em um único arquivo de saída, realizando limpeza e transformações inteligentes no processo.</p>
                <p>Navegue pelos tópicos à esquerda para aprender sobre cada funcionalidade.</p>
            """,
            "1. Seleção e Leitura": """
                <h2>1. Seleção de Pasta e Opções de Leitura</h2>
                <p><b>Selecionar Pasta:</b> O primeiro passo é sempre selecionar a pasta onde seus arquivos de dados estão localizados.</p>
                <p><b>Atualizar Pasta:</b> Se você adicionar ou remover arquivos da pasta com o programa aberto, ou também alterar o delimitador, clique no botão 'Atualizar' (com o ícone de recarregar) para que a lista de arquivos e delimitador sejam atualizados.</p>
                <p><b>Delimitador:</b> Para arquivos <b>.CSV</b> e <b>.TXT</b>, é crucial escolher o caractere que separa as colunas (delimitador). As opções mais comuns estão disponíveis, ou você pode especificar um customizado em 'Outro...'.</p>
                <br>
                <h3>Detecção Automática de Cabeçalho</h3>
                <p>A ferramenta detecta automaticamente em qual linha o cabeçalho se encontra, ignorando títulos ou linhas em branco no topo dos arquivos. Isso funciona tanto na pré-visualização quanto na consolidação final.</p>
            """,
            "2. Mapeamento de Cabeçalhos": """
                <h2>2. Mapeamento e Agrupamento Inteligente</h2>
                <p>Esta é a etapa mais poderosa da ferramenta.</p>
                <p><b>Análise Inteligente:</b> Ao clicar em 'Analisar/Mapear Cabeçalhos', o programa lê uma amostra de cada arquivo e cria uma "impressão digital" de cada coluna, analisando não apenas o nome, mas também o tipo de dado do conteúdo.</p>
                <p><b>Grupos Sugeridos:</b> Com base nessa análise, ele agrupa automaticamente colunas que parecem ser a mesma coisa, mesmo que tenham nomes diferentes (ex: 'CNPJ' e 'C.N.P.J.').</p>
                <p><b>Tooltip de Detalhes:</b> Passe o mouse sobre um nome na coluna 'Grupo Sugerido' para ver todas as variações originais que foram agrupadas ali.</p>
                <p><b>Sua Supervisão:</b> Você tem o controle final! Use os botões <b>'Dividir Grupo'</b> para separar colunas agrupadas incorretamente, e <b>'Mesclar Grupos'</b> para unir grupos que você sabe que são a mesma coisa.</p>
            """,
            "3. Filtros de Dados": """
                <h2>3. Filtros de Dados</h2>
                <p>A funcionalidade de filtro permite refinar os dados que serão incluídos na consolidação final.</p>
                <p><b>Lógica E / OU:</b> O sistema aplica uma lógica inteligente:</p>
                <ul>
                    <li>Regras para a <b>mesma coluna</b> são combinadas com <b>OU</b> (ex: `Nome = "João" OU Nome = "Maria"`).</li>
                    <li>Regras para <b>colunas diferentes</b> são combinadas com <b>E</b> (ex: `(Nome = "João") E (Status = "Ativo")`).</li>
                    <li>Regras de exclusão ('Não contém', 'Diferente de') são aplicadas após as de inclusão.</li>
                </ul>
                <p>Isso permite criar filtros complexos, como 'incluir todas as Compras e Vendas, mas não incluir as que forem Vendas de Ativos'.</p>
            """,
            "4. Consolidação e Saída": """
                <h2>4. Consolidação e Arquivo Final</h2>
                <p>Após definir o mapeamento e os filtros, escolha o local e o formato de saída (XLSX ou CSV) e clique em 'Iniciar Consolidação'.</p>
                <h3>Recursos da Saída em Excel</h3>
                <ul>
                    <li><b>Coluna 'Origem':</b> Uma coluna extra é adicionada ao final, indicando de qual arquivo e aba cada linha de dado veio.</li>
                    <li><b>Estilo Profissional:</b> O cabeçalho é formatado com um fundo escuro e texto claro.</li>
                    <li><b>Múltiplas Abas:</b> Se o resultado tiver mais de ~1 milhão de linhas, ele será automaticamente dividido em múltiplas abas ('Dados_Parte_1', 'Dados_Parte_2', etc.).</li>
                    <li><b>Formatação Adicional:</b> A planilha vem com painéis congelados, zoom ajustado e sem linhas de grade para uma melhor visualização.</li>
                </ul>
            """,
        }

        # --- Layout da Janela ---
        layout = QHBoxLayout(self)

        # Painel esquerdo com os tópicos
        self.topics_list = QListWidget()
        self.topics_list.setMaximumWidth(200)
        self.topics_list.addItems(self.help_content.keys())
        
        # Painel direito para exibir o conteúdo
        self.content_display = QTextEdit()
        self.content_display.setReadOnly(True)

        layout.addWidget(self.topics_list)
        layout.addWidget(self.content_display)
        
        # Conectar a seleção de tópico à exibição de conteúdo
        self.topics_list.currentItemChanged.connect(self.display_topic_content)
        # Selecionar o primeiro item por padrão
        self.topics_list.setCurrentRow(0)

    def display_topic_content(self, current_item, previous_item):
        """Exibe o conteúdo HTML do tópico selecionado."""
        if current_item:
            topic = current_item.text()
            self.content_display.setHtml(self.help_content.get(topic, "<p>Tópico não encontrado.</p>"))

class SheetSelectionDialog(QDialog):
    """Diálogo para seleção global de abas."""
    def __init__(self, unique_sheet_names, existing_rules=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Seleção Global de Abas")
        self.setMinimumSize(500, 450)

        layout = QVBoxLayout(self)
        
        # --- Modo de Seleção (Incluir vs Excluir) ---
        mode_box = QGroupBox("Modo de Operação")
        mode_layout = QVBoxLayout()
        self.include_radio = QRadioButton("Incluir APENAS as abas marcadas abaixo")
        self.exclude_radio = QRadioButton("Excluir TODAS as abas marcadas abaixo")
        mode_layout.addWidget(self.include_radio)
        mode_layout.addWidget(self.exclude_radio)
        mode_box.setLayout(mode_layout)
        layout.addWidget(mode_box)
        
        # --- Lista de Abas ---
        list_box = QGroupBox("Abas Encontradas em Todos os Arquivos")
        list_layout = QVBoxLayout()
        
        self.list_widget = QListWidget()
        for name in sorted(list(unique_sheet_names)):
            item = QListWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.list_widget.addItem(item)
        
        list_layout.addWidget(self.list_widget)

        # Botões de ação para a lista
        list_actions_layout = QHBoxLayout()
        mark_all_button = QPushButton("Marcar Todos")
        unmark_all_button = QPushButton("Desmarcar Todos")
        list_actions_layout.addStretch()
        list_actions_layout.addWidget(mark_all_button)
        list_actions_layout.addWidget(unmark_all_button)
        list_layout.addLayout(list_actions_layout)
        
        list_box.setLayout(list_layout)
        layout.addWidget(list_box)

        # --- Botões OK e Cancelar ---
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel | QDialogButtonBox.Reset)
        button_box.button(QDialogButtonBox.Reset).setText("Limpar Regras")
        layout.addWidget(button_box)

        # --- Conectar Sinais ---
        mark_all_button.clicked.connect(lambda: self._set_all_check_state(Qt.Checked))
        unmark_all_button.clicked.connect(lambda: self._set_all_check_state(Qt.Unchecked))
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        button_box.button(QDialogButtonBox.Reset).clicked.connect(self.clear_rules)

        # --- Carregar regras existentes ---
        if existing_rules and existing_rules.get("names"):
            if existing_rules.get("mode") == "exclude":
                self.exclude_radio.setChecked(True)
            else:
                self.include_radio.setChecked(True)
            
            for i in range(self.list_widget.count()):
                item = self.list_widget.item(i)
                if item.text() in existing_rules["names"]:
                    item.setCheckState(Qt.Checked)
        else:
            # Estado padrão
            self.include_radio.setChecked(True)

    def _set_all_check_state(self, state):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(state)

    def clear_rules(self):
        """Limpa a seleção e reseta para o modo padrão."""
        self.include_radio.setChecked(True)
        self._set_all_check_state(Qt.Unchecked)
        # Opcional: pode fechar o diálogo ou apenas limpar a tela.
        # Por enquanto, apenas limpa a seleção.

    def get_rules(self):
        """Retorna as regras definidas pelo usuário."""
        mode = "exclude" if self.exclude_radio.isChecked() else "include"
        selected_names = set()
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item.checkState() == Qt.Checked:
                selected_names.add(item.text())
        
        # Se nenhuma aba for marcada, retorna uma regra vazia para desativar o filtro global
        if not selected_names:
            return {}
            
        return {"mode": mode, "names": selected_names}

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon("app/logo1.png"))
        self.header_mapping = {}
        self.filter_rules = []
        self.pivot_rules = {}
        self.duplicate_key_columns = []
        self.sheet_selection_rules = {}
        self.all_sheets_cache = {}
        menu_bar = self.menuBar()
        
        # Menu "Ajuda"
        help_menu = menu_bar.addMenu("&Ajuda") # O & cria um atalho (Alt+A)
        
        guide_action = QAction("Guia do Usuário...", self)
        guide_action.triggered.connect(self.open_help_dialog)
        help_menu.addAction(guide_action)
        self.setWindowTitle("DataFlow")
        self.setGeometry(100, 100, 1000, 750) 

        self.current_files_paths = {}
        self.output_file_path = "" 
        self.consolidation_thread = None
        self.sheet_loader_thread = None
        self.header_analyzer_thread = None
        self.sheet_analysis_worker = None
        self.is_last_log_progress = False
        self.sheet_selections = {} 
        self.last_used_input_folder = self._load_last_input_folder() # <--- CARREGAR AO INICIAR

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # --- 1. Seção de Seleção de Pasta ---
        folder_selection_layout = QHBoxLayout()
        # logo_label = QLabel()
        # pixmap = QPixmap("app/logo.png")
        # logo_label.setPixmap(pixmap.scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        logo_widget = QSvgWidget("app/logo2.svg")
        logo_widget.setFixedSize(75, 75)
        folder_selection_layout.addWidget(logo_widget)
        self.folder_path_label = QLabel("Pasta do Projeto:")
        self.folder_path_line_edit = QLineEdit()
        self.folder_path_line_edit.setReadOnly(True)
        # Se uma pasta foi carregada, exibi-la (opcional, ou só usar no diálogo)
        if self.last_used_input_folder:
            self.folder_path_line_edit.setText(self.last_used_input_folder)
            # Poderia até chamar list_files_in_folder aqui se quisesse carregar automaticamente
            # self.list_files_in_folder(self.last_used_input_folder) 
            # Mas vamos manter o clique do usuário para carregar os arquivos.
        self.select_folder_button = QPushButton("Selecionar Pasta...")
        folder_icon = self.style().standardIcon(QStyle.SP_DirIcon)
        self.select_folder_button.setIcon(folder_icon)
        self.select_folder_button.clicked.connect(self.open_folder_dialog)

        # BOTÃO para atualizar
        self.refresh_button = QPushButton("Atualizar")
        refresh_icon = self.style().standardIcon(QStyle.SP_BrowserReload)
        self.refresh_button.setIcon(refresh_icon)
        self.refresh_button.clicked.connect(self.refresh_folder_list)
        self.refresh_button.setEnabled(False)
        self.refresh_button.setToolTip("Atualiza a lista de arquivos da pasta selecionada.")

        # BOTÃO para mapear cabeçalhos
        self.map_headers_button = QPushButton("Analisar/Mapear Cabeçalhos")
        header_icon = self.style().standardIcon(QStyle.SP_FileDialogContentsView)
        self.map_headers_button.setIcon(header_icon)
        self.define_filters_button = QPushButton("Definir Filtros")
        filter_icon = self.style().standardIcon(QStyle.SP_DialogHelpButton)
        self.define_filters_button.setIcon(filter_icon)
        self.define_filters_button.clicked.connect(self.open_filter_dialog)
        self.define_filters_button.setEnabled(False) # Habilitar somente após mapeamento de cabeçalho
        self.define_filters_button.setToolTip("Filtre os valores das colunas e remova dados desnecessários.")
        self.map_headers_button.clicked.connect(self.open_header_mapping_dialog)
        self.map_headers_button.setEnabled(False) # Habilitar após selecionar pasta 
        self.map_headers_button.setToolTip("Agrupe cabeçalhos, analise seus tipos de dados, remova duplicatas e escolha quais serão selecionados.")
        self.sheet_selection_button = QPushButton("Seleção de Abas...")
        sheet_icon = self.style().standardIcon(QStyle.SP_MessageBoxInformation)
        self.sheet_selection_button.setIcon(sheet_icon)
        self.sheet_selection_button.clicked.connect(self.open_sheet_selection_dialog)
        self.sheet_selection_button.setEnabled(False) # Começa desabilitado
        self.sheet_selection_button.setToolTip("Defina regras globais para incluir ou excluir abas em todos os arquivos Excel.")

        self.pivot_button = QPushButton("Criar Tabela de Resumo...")
        pivot_icon = self.style().standardIcon(QStyle.SP_FileDialogEnd)
        self.pivot_button.setIcon(pivot_icon)
        self.pivot_button.clicked.connect(self.open_pivot_dialog)
        self.pivot_button.setEnabled(False)
        self.pivot_button.setToolTip("Cria uma tabela resumo (dinâmica) a partir dos dados consolidados.")
        
        folder_selection_layout.addWidget(logo_widget)
        folder_selection_layout.addWidget(self.folder_path_label)
        folder_selection_layout.addWidget(self.folder_path_line_edit)
        folder_selection_layout.addWidget(self.select_folder_button)
        folder_selection_layout.addWidget(self.refresh_button)
        main_layout.addLayout(folder_selection_layout)

        config_buttons_layout = QHBoxLayout()   
        config_buttons_layout.setContentsMargins(0, 10, 0, 0) # Adiciona um espaçamento superior
        config_buttons_layout.addWidget(self.map_headers_button)
        config_buttons_layout.addWidget(self.sheet_selection_button)
        config_buttons_layout.addWidget(self.define_filters_button)
        config_buttons_layout.addWidget(self.pivot_button)
        config_buttons_layout.addStretch() # Empurra os botões para a esquerda
        main_layout.addLayout(config_buttons_layout)
        # --- Opções de Leitura ---
        self.options_group_box = QGroupBox("Opções de Leitura para .CSV e .TXT")
        options_layout = QHBoxLayout()
        delimiter_label = QLabel("Delimitador: ")
        self.delimiter_combo = QComboBox()
        self.delimiter_combo.addItems([
            "Ponto e Vírgula (;)",
            "Vírgula (,)",
            "Tabulação (Tab)",
            "Pipe (|)",
            "Outro..."
        ])
        self.delimiter_custom_edit = QLineEdit()
        self.delimiter_custom_edit.setPlaceholderText("Digite o delimitador")
        self.delimiter_custom_edit.setFixedWidth(120)
        self.delimiter_custom_edit.setVisible(False) # Começa oculto
        self.delimiter_combo.currentTextChanged.connect(self._on_delimiter_changed)
        options_layout.addWidget(delimiter_label)
        options_layout.addWidget(self.delimiter_combo)
        options_layout.addWidget(self.delimiter_custom_edit)
        options_layout.addStretch() # Empurra tudo para a esquerda
        self.options_group_box.setLayout(options_layout)
        main_layout.addWidget(self.options_group_box)

        # --- 2. Seção Intermediária (Arquivos/Abas e Console) ---
        middle_section_layout = QHBoxLayout()
        left_panel_layout = QVBoxLayout()
        
        self.files_label = QLabel("Arquivos Encontrados (.xlsx, .csv, .xls, .txt):")
        self.files_list_widget = QListWidget()
        self.files_list_widget.currentItemChanged.connect(self.on_file_selected_for_preview)
        
        self.sheets_label = QLabel("Abas do Arquivo Excel Selecionado (marque para incluir):")
        self.sheets_list_widget = QListWidget()
        self.sheets_list_widget.setEnabled(False) 
        # Conectar o sinal itemChanged para quando o estado de um checkbox de aba mudar
        self.sheets_list_widget.itemChanged.connect(self.on_sheet_selection_changed)
        self.sheets_list_widget.currentItemChanged.connect(self.on_sheet_list_item_selected_for_preview)

        left_panel_layout.addWidget(self.files_label)
        left_panel_layout.addWidget(self.files_list_widget)
        left_panel_layout.addWidget(self.sheets_label)
        left_panel_layout.addWidget(self.sheets_list_widget)
        # --- Botões para Marcar/Desmarcar Abas ---
        sheet_actions_layout = QHBoxLayout()
        self.mark_all_sheets_button = QPushButton("Marcar Todas")
        self.mark_all_sheets_button.clicked.connect(self.mark_all_sheets)
        self.unmark_all_sheets_button = QPushButton("Desmarcar Todas")
        self.unmark_all_sheets_button.clicked.connect(self.unmark_all_sheets)
        
        sheet_actions_layout.addWidget(self.mark_all_sheets_button)
        sheet_actions_layout.addWidget(self.unmark_all_sheets_button)
        left_panel_layout.addLayout(sheet_actions_layout)
        
        self.mark_all_sheets_button.setEnabled(False)
        self.unmark_all_sheets_button.setEnabled(False)

        middle_section_layout.addLayout(left_panel_layout)

        # 2.2 Painel Direito (Abas para Console e Pré-visualização)
        right_tab_widget = QTabWidget() 

        # Aba do Console de Log
        log_console_widget = QWidget()
        log_layout = QVBoxLayout(log_console_widget)
        self.log_label = QLabel("Console de Log:") # Pode ser removido se o título da aba for suficiente
        self.log_console_text_edit = QTextEdit()
        self.log_console_text_edit.setReadOnly(True)
        
        log_layout.addWidget(self.log_console_text_edit) # Adiciona diretamente, sem label se preferir
        right_tab_widget.addTab(log_console_widget, "Console de Log")

        # Aba de Pré-visualização de Dados
        preview_widget = QWidget()
        preview_layout = QVBoxLayout(preview_widget)
        self.preview_table_view = QTableView()
        self.preview_table_model = PolarsTableModel() # Instancia nosso modelo customizado
        self.preview_table_view.setModel(self.preview_table_model)
        # self.preview_table_view.setEditTriggers(QTableView.NoEditTriggers) # Desabilitar edição
        self.preview_table_view.setAlternatingRowColors(True)
        preview_layout.addWidget(self.preview_table_view)
        right_tab_widget.addTab(preview_widget, "Pré-visualização de Dados")
        
        middle_section_layout.addWidget(right_tab_widget) # Adiciona o QTabWidget ao layout
        
        middle_section_layout.setStretchFactor(left_panel_layout, 1) 
        middle_section_layout.setStretchFactor(right_tab_widget, 3) # Dar mais espaço para o painel direito
        main_layout.addLayout(middle_section_layout)

        # --- 3. Seção de Configuração de Saída ---
        output_config_layout = QHBoxLayout()
        self.output_name_label = QLabel("Nome do Arquivo de Saída:")
        self.output_name_line_edit = QLineEdit("consolidado") # Nome padrão
        
        self.output_format_label = QLabel("Formato:")
        self.output_format_combo_box = QComboBox()
        self.output_format_combo_box.addItems(["XLSX", "CSV", "Parquet"])
        self.output_format_combo_box.currentTextChanged.connect(self.update_output_filename_extension)
        
        self.save_as_button = QPushButton("Salvar Como...")
        self.save_as_button.clicked.connect(self.open_save_file_dialog)

        output_config_layout.addWidget(self.output_name_label)
        output_config_layout.addWidget(self.output_name_line_edit)
        output_config_layout.addWidget(self.output_format_label)
        output_config_layout.addWidget(self.output_format_combo_box)
        output_config_layout.addWidget(self.save_as_button)
        main_layout.addLayout(output_config_layout)

        # --- 4. Seção de Ação e Progresso ---
        action_progress_layout = QVBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)

        # self.progress_text_label = QLabel("")
        # self.progress_text_label.setAlignment(Qt.AlignCenter)
        # self.progress_text_label.setVisible(False)

        self.consolidate_button = QPushButton("Iniciar Consolidação")
        self.consolidate_button.setObjectName("consolidate_button")
        self.consolidate_button.clicked.connect(self.start_consolidation) 

        self.cancel_button = QPushButton("Cancelar")
        self.cancel_button.clicked.connect(self.cancel_consolidation)
        self.cancel_button.setVisible(False) 

        buttons_layout = QHBoxLayout() 
        buttons_layout.addWidget(self.consolidate_button)
        buttons_layout.addWidget(self.cancel_button)
        
        # action_progress_layout.addWidget(self.progress_text_label)
        action_progress_layout.addWidget(self.progress_bar)
        action_progress_layout.addLayout(buttons_layout) 
        main_layout.addLayout(action_progress_layout)

        main_layout.setSpacing(15)
        self.show()
        self.log_message("Aplicação iniciada. Selecione uma pasta para começar.", LogLevel.INFO)
        self.update_output_filename_extension(self.output_format_combo_box.currentText())
    
    def open_pivot_dialog(self):
        """Abre o diálogo de configuração da tabela de resumo."""
        if not self.header_mapping:
            self.log_message("Por favor, analise e mapeie os cabeçalhos primeiro.", LogLevel.WARNING)
            return

        # Coleta os nomes e tipos das colunas finais
        final_headers_info = {}
        for map_info in self.header_mapping.values():
            if map_info.get('include'):
                final_name = map_info['final_name']
                type_str = map_info['type_str']
                final_headers_info[final_name] = type_str
        
        all_final_headers = sorted(list(final_headers_info.keys()))
        numeric_types = {"Inteiro", "Decimal (Float)"}
        numeric_headers = sorted([h for h, t in final_headers_info.items() if t in numeric_types])

        if not numeric_headers:
            self.log_message("Nenhuma coluna numérica encontrada no mapeamento para usar nos cálculos.", LogLevel.WARNING)

        dialog = PivotDialog(all_final_headers, numeric_headers, self.pivot_rules, self)
        if dialog.exec() == QDialog.Accepted:
            self.pivot_rules = dialog.get_rules()
            if self.pivot_rules:
                self.log_message("Regras da tabela de resumo foram definidas.", LogLevel.SUCCESS)
            else:
                self.log_message("Regras da tabela de resumo foram limpas.", LogLevel.INFO)

    def open_sheet_selection_dialog(self):
        """Inicia a análise de todas as abas em uma thread e abre o diálogo de seleção ao concluir."""
        if self.sheet_analysis_worker and self.sheet_analysis_worker.isRunning():
            self.log_message("A análise de abas já está em andamento.", LogLevel.INFO)
            return

        excel_files = [path for path in self.current_files_paths.values() if path.lower().endswith((".xlsx", ".xls"))]
        if not excel_files:
            self.log_message("Nenhum arquivo Excel encontrado na pasta para analisar as abas.", LogLevel.WARNING)
            return

        self.log_message("Analisando nomes de todas as abas dos arquivos Excel (em segundo plano)...", LogLevel.INFO)
        self.sheet_selection_button.setEnabled(False) # Desabilita durante a análise

        self.sheet_analysis_worker = SheetAnalysisWorker(excel_files)
        self.sheet_analysis_worker.finished.connect(self.on_sheet_analysis_finished)
        self.sheet_analysis_worker.start()

    def on_sheet_analysis_finished(self, all_sheets_cache, unique_sheet_names, error_message):
        """Chamado quando a SheetAnalysisWorker termina."""
        self.sheet_selection_button.setEnabled(True) # Reabilita o botão
        self.sheet_analysis_worker = None

        if error_message:
            self.log_message(f"Erro durante a análise de abas: {error_message}", LogLevel.ERROR)
            return

        self.log_message(f"Análise de abas concluída. {len(unique_sheet_names)} nomes de abas únicos encontrados.", LogLevel.SUCCESS)
        self.all_sheets_cache = all_sheets_cache # Salva o cache para uso posterior

        # Abre o diálogo com as abas encontradas e as regras que já possam existir
        dialog = SheetSelectionDialog(unique_sheet_names, self.sheet_selection_rules, self)
        if dialog.exec() == QDialog.Accepted:
            self.sheet_selection_rules = dialog.get_rules()
            if self.sheet_selection_rules:
                self.log_message(f"Regra de seleção de abas definida. Modo: {self.sheet_selection_rules['mode']}.", LogLevel.SUCCESS)
            else:
                self.log_message("Regras de seleção global de abas foram limpas.", LogLevel.INFO)

    def open_help_dialog(self):
        """Cria e exibe a janela de ajuda/guia do usuário."""
        # O diálogo já tem o texto, não precisamos passar nada.
        # Passamos 'self' para que o diálogo seja "filho" da janela principal.
        dialog = HelpDialog(self)
        dialog.exec() # .exec() abre o diálogo de forma moda
    
    def _on_delimiter_changed(self, text):
        is_custom = (text == "Outro...")
        self.delimiter_custom_edit.setVisible(is_custom)
    
    def refresh_folder_list(self):
        """Recarrega a lista de arquivos da pasta
        atualmente exibida."""
        current_folder = self.folder_path_line_edit.text()
        if not os.path.isdir(current_folder):
            self.log_message("Nenhuma pasta válida selecionada para atualizar.", LogLevel.WARNING)
            return
        self.log_message(f"Atualizando lista de arquivos para: {current_folder}", LogLevel.INFO)
        self.list_files_in_folder(current_folder)

    def get_selected_delimiter(self):
        try:
            selected = self.delimiter_combo.currentText()
            if selected == "Outro...":
                return self.delimiter_custom_edit.text()
            elif selected == "Tabulação (Tab)":
                return '\t'
            elif "(" in selected and ")" in selected:
                return selected.split("(")[1].replace(")", "")
            else:
                return selected
        except RuntimeError as e:
            if "already deleted" in str(e):
                self.log_message("AVISO: Widget de delimitador foi destruído inesperadamente. Usando ';' como padrão.", LogLevel.WARNING)
                return "|"
            else:
                raise
    
    def mark_all_sheets(self):
        self._set_all_sheets_check_state(Qt.Checked)
    
    def unmark_all_sheets(self):
        self._set_all_sheets_check_state(Qt.Unchecked)
    
    def _set_all_sheets_check_state(self, check_state):
        if not self.sheets_list_widget.isEnabled() or self.sheets_list_widget.count() == 0:
            return
        try:
            self.sheets_list_widget.itemChanged.disconnect(self.on_sheet_selection_changed)
        except RuntimeError:
            pass
        for i in range(self.sheets_list_widget.count()):
            item = self.sheets_list_widget.item(i)
            item.setCheckState(check_state)
            current_file_item = self.files_list_widget.currentItem()
            if current_file_item:
                file_path = self.current_files_paths.get(current_file_item.text())
                if file_path and file_path in self.sheet_selections:
                    self.sheet_selections[file_path][item.text()] = (check_state == Qt.Checked)
        self.sheets_list_widget.itemChanged.connect(self.on_sheet_selection_changed)
        action = "marcadas" if check_state == Qt.Checked else "desmarcadas"
        self.log_message(f'Todas as abas visíveis foram {action}.', LogLevel.INFO)

    def _get_config_path(self):
        """Retorna o caminho para o arquivo de configuração da aplicação."""
        # Salvar na pasta do usuário (mais robusto) ou na pasta da aplicação
        # Usar AppData ou .config no Linux/macOS é o ideal, mas para simplicidade:
        try:
            # Tenta obter o diretório do script
            base_path = os.path.dirname(os.path.abspath(sys.argv[0]))
        except:
            # Fallback para o diretório de trabalho atual se sys.argv[0] não for confiável (ex: PyInstaller one-file)
            base_path = os.getcwd() 
        return os.path.join(base_path, CONFIG_FILE_NAME)

    def _load_last_input_folder(self):
        """Carrega o último caminho da pasta de entrada do arquivo de configuração."""
        config_path = self._get_config_path()
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='latin-1') as f:
                    config_data = json.load(f)
                    return config_data.get("last_input_folder")
        except Exception as e:
            # Não precisa ser um erro crítico, apenas logar um aviso
            print(f"Aviso: Não foi possível carregar a configuração: {e}") # Usar print para log antes do logger da GUI estar pronto
        return None

    def _save_last_input_folder(self, folder_path):
        """Salva o caminho da pasta de entrada no arquivo de configuração."""
        config_path = self._get_config_path()
        config_data = {"last_input_folder": folder_path}
        try:
            with open(config_path, 'w', encoding='latin-1') as f:
                json.dump(config_data, f, indent=4)
            # Não precisa logar na GUI cada vez que salva, a menos que queira
            # self.log_message("Diretório padrão salvo.", LogLevel.INFO) 
        except Exception as e:
            self.log_message(f"Erro ao salvar configuração de diretório: {e}", LogLevel.ERROR)


    def open_folder_dialog(self):
        # Usar o último diretório salvo ou o diretório do QLineEdit ou o home do usuário
        start_dir = self.last_used_input_folder or self.folder_path_line_edit.text() or os.path.expanduser("~")
        
        folder_path = QFileDialog.getExistingDirectory(self, "Selecionar Pasta do Projeto", start_dir)
        
        if folder_path:
            self.folder_path_line_edit.setText(folder_path)
            self.log_message(f"Pasta selecionada: {folder_path}", LogLevel.INFO)
            self.list_files_in_folder(folder_path)
            self.last_used_input_folder = folder_path # Atualizar para o próximo uso
            self._save_last_input_folder(folder_path) # <--- SALVAR AO SELECIONAR
        else:
            self.log_message("Seleção de pasta cancelada.", LogLevel.INFO)
    
    # --- Métodos para Mapeamento de Cabeçalhos ---
    def open_header_mapping_dialog(self):
        """Coleta cabeçalhos e abre o diálogo de mapeamento."""
        if not self.current_files_paths:
            self.log_message("Selecione uma pasta e arquivos primeiro.", LogLevel.WARNING)
            return
        
        if self.header_analyzer_thread and self.header_analyzer_thread.isRunning():
            self.log_message("A análise de cabeçalhos já está em andamento.", LogLevel.INFO)
            return # Evita iniciar múltiplas análises

        # self.log_message("Analisando cabeçalhos dos arquivos selecionados...", LogLevel.INFO)
        # unique_headers = set()
        
        # Usar get_files_and_sheets_to_process para saber quais arquivos/abas considerar
        files_and_sheets_config = self.get_files_and_sheets_to_process()
        if not files_and_sheets_config:
            self.log_message("Nenhum arquivo/aba configurado para processamento. Não é possível mapear cabeçalhos.", LogLevel.WARNING)
            return
        
        self.log_message("Analisando cabeçalhos dos arquivos selecionados (em segundo plano)...", LogLevel.INFO)
        self.map_headers_button.setEnabled(False) # Desabilita botão para evitar cliques duplos
        selected_delimiter = self.get_selected_delimiter()
        if not selected_delimiter:
            self.log_message("Análise de cabeçalhos falhou: Delimitador inválido para CSV/TXT.", LogLevel.ERROR)
            self.map_headers_button.setEnabled(True)
            return
        self.define_filters_button.setEnabled(False)
        self.pivot_button.setEnabled(False)

        self.filter_rules.clear()
        self.header_analyzer_thread = HeaderAnalysisWorker(files_and_sheets_config, selected_delimiter)
        self.header_analyzer_thread.finished.connect(self.on_header_analysis_finished)
        self.header_analyzer_thread.progress_log.connect(self.log_message)
        self.header_analyzer_thread.start()

        # =============== DESCARTADA ===============
        # Pequena thread para não travar a GUI ao ler cabeçalhos
        # (Pode ser excessivo para apenas cabeçalhos, mas bom se houver muitos arquivos)
        # Por simplicidade no MVP do mapeamento, vamos fazer síncrono por enquanto.
        # Se ficar lento, podemos mover para uma thread.
        
        # processed_for_headers = 0
        # for file_path, selected_sheets in files_and_sheets_config:
        #     file_name = os.path.basename(file_path)
        #     try:
        #         if file_path.lower().endswith(".csv"):
        #             reader = pl.read_csv_batched(file_path, batch_size=5, infer_schema_length=0, encoding = 'latin-1', separator = ';') 
        #             batches = reader.next_batches(1) 
                    
        #             if batches and len(batches) > 0:
        #                     first_batch = batches[0]
        #                     if first_batch is not None and first_batch.width > 0:
        #                         unique_headers.update(first_batch.columns)
        #             # else: Tratar CSV vazio ou só com cabeçalho (opcional)

        #         elif file_path.lower().endswith((".xlsx", ".xls")) and selected_sheets:
        #             for sheet_name in selected_sheets:
        #                 # Ler apenas a primeira linha da aba
        #                 # df_header = pl.read_excel(file_path, sheet_name=sheet_name, n_rows=1, infer_schema_length=0) # n_rows não é para read_excel
        #                 # Ler e pegar head(0).columns ou head(1).columns
        #                 try:
        #                     df_sample = pl.read_excel(file_path, sheet_name=sheet_name).head(0) # Pega só cabeçalhos
        #                 except pl.exceptions.PolarsError as e_excel:
        #                     error_msg_lower = str(e_excel).lower()
        #                     is_numeric_conversion_error = ('conversion' in error_msg_lower and 'f64' in error_msg_lower or 'float' in error_msg_lower) and ('i64' in error_msg_lower or 'int' in error_msg_lower) and 'failed' in error_msg_lower
        #                     if is_numeric_conversion_error:
        #                         try:
        #                             df_temp = pl.read_excel(source = file_path, sheet_name = sheet_name, infer_schema_length = 0)
        #                             if df_temp is not None:
        #                                 df_sample = df_temp.select([pl.all().cast(pl.String)])
        #                                 # df_preview_sliced = df_original.head(n_rows_to_preview)
        #                             else:
        #                                 df_sample = None
        #                         except Exception as e_retry_excel:
        #                             df_sample = None
        #                     else:
        #                         df_sample = None
        #                 if df_sample.width > 0:
        #                      unique_headers.update(df_sample.columns)
        #         processed_for_headers +=1
        #         # TODO: Adicionar feedback de progresso se for demorado
        #     except Exception as e:
        #         self.log_message(f"Erro ao ler cabeçalhos de '{file_name}' (Aba: {sheet_name if selected_sheets else 'N/A'}): {e}", LogLevel.ERROR)

        # if not unique_headers:
        #     self.log_message("Nenhum cabeçalho encontrado ou erro ao ler todos os cabeçalhos.", LogLevel.WARNING)
        #     return

        # self.log_message(f"Cabeçalhos únicos encontrados: {len(unique_headers)}", LogLevel.SUCCESS)
        
        # # Passar o self.header_mapping existente para o diálogo
        # dialog = HeaderMappingDialog(unique_headers, self, self.header_mapping)
        # if dialog.exec() == QDialog.Accepted: # .exec_() em PyQt5
        #     self.header_mapping = dialog.get_mapping()
        #     self.log_message("Mapeamento de cabeçalhos atualizado.", LogLevel.SUCCESS)
        #     # Logar o mapeamento para depuração (opcional)
        #     # for original, map_info in self.header_mapping.items():
        #     #     self.log_message(f"  '{original}' -> '{map_info['final_name']}' (Incluir: {map_info['include']})", LogLevel.INFO)
        # else:
        #     self.log_message("Mapeamento de cabeçalhos cancelado.", LogLevel.INFO)
    
    def on_header_analysis_finished(self, suggested_groups, error_object):
        """Chamado quando a HeaderAnalysisWorker termina."""
        self.map_headers_button.setEnabled(True) # Reabilita o botão
        if self.header_analyzer_thread: # Garante que a thread exista antes de tentar limpá-la
            self.header_analyzer_thread = None # Limpa a referência da thread

        if error_object:
            if isinstance(error_object, InterruptedError):
                self.log_message(f"Análise de cabeçalhos cancelada: {error_object}", LogLevel.WARNING)
            else:
                self.log_message(f"Erro durante a análise de cabeçalhos: {error_object}", LogLevel.ERROR)
            return

        if not suggested_groups:
            self.log_message("Nenhum cabeçalho encontrado ou erro ao ler todos os cabeçalhos. Verifique os arquivos e seleções de abas.", LogLevel.WARNING)
            return

        self.log_message(f"Análise concluída. Cabeçalhos únicos encontrados: {len(suggested_groups)}", LogLevel.SUCCESS)

        # Passar o self.header_mapping existente para o diálogo
        dialog = HeaderMappingDialog(suggested_groups, self, self.header_mapping, self.duplicate_key_columns)
        if dialog.exec() == QDialog.Accepted:
            self.header_mapping = dialog.get_mapping()
            # --- Salva as colunas para checagem de duplicatas ---
            self.duplicates_config = dialog.get_duplicates_config()
            # self.duplicate_key_columns = dialog.get_duplicate_check_columns()
            self.log_message("Mapeamento de cabeçalhos atualizado.", LogLevel.SUCCESS)
            key_columns = self.duplicates_config.get("key_columns", [])
            if key_columns:
                report_msg = "e um relatório sera gerado" if self.duplicates_config.get("generate_report") else ""
                self.log_message(f"Remoção de duplicatas ativada para as chaves: {', '.join(key_columns)} {report_msg}", LogLevel.INFO)
            self.define_filters_button.setEnabled(True)
            self.pivot_button.setEnabled(True)
            # Opcional: Logar o mapeamento para depuração
            # for original, map_info in self.header_mapping.items():
            #     self.log_message(f"  '{original}' -> '{map_info['final_name']}' (Tipo: {map_info['type_str']}, Incluir: {map_info['include']})", LogLevel.INFO)
        else:
            self.log_message("Mapeamento de cabeçalhos cancelado.", LogLevel.INFO)
    
    def on_file_selected_for_preview(self, current_file_item, previous_file_item):
        """Chamado quando um ARQUIVO é selecionado na lista.
           Também chama o antigo on_file_selected para carregar as abas e suas seleções.
        """
        self.on_file_selected(current_file_item, previous_file_item) # Chama a lógica existente de abas
        
        # Limpar pré-visualização se nenhum item ou arquivo não Excel/CSV
        if not current_file_item:
            self.preview_table_model.clear_data()
            return

        file_name = current_file_item.text()
        file_path = self.current_files_paths.get(file_name)

        if not file_path:
            self.preview_table_model.clear_data()
            return

        if file_path.lower().endswith((".csv", ".txt")):
            self.update_preview(file_path) # Pré-visualiza CSV diretamente
        elif file_path.lower().endswith((".xlsx", ".xls")):
            # Para Excel, a pré-visualização da aba será acionada por:
            # a) on_sheet_loading_finished (que seleciona a primeira aba) -> on_sheet_list_item_selected_for_preview
            # b) clique manual do usuário em uma aba -> on_sheet_list_item_selected_for_preview
            # Podemos limpar a pré-visualização aqui, e ela será preenchida quando as abas carregarem e uma for selecionada.
            self.preview_table_model.clear_data()
        else: # Outros tipos de arquivo (não deve acontecer se o filtro de arquivos estiver correto)
            self.preview_table_model.clear_data()


    def on_sheet_list_item_selected_for_preview(self, current_sheet_item, previous_sheet_item):
        """Chamado quando uma ABA é selecionada na lista de abas."""
        if not current_sheet_item:
            self.preview_table_model.clear_data()
            return

        current_file_item = self.files_list_widget.currentItem()
        if not current_file_item:
            self.preview_table_model.clear_data()
            return

        file_name = current_file_item.text()
        file_path = self.current_files_paths.get(file_name)
        sheet_name = current_sheet_item.text()

        if file_path and sheet_name:
            self.update_preview(file_path, sheet_name)
        else:
            self.preview_table_model.clear_data()


    def update_preview(self, file_path, sheet_name=None, n_rows_to_preview=50):
        self.log_message(f"Gerando pré-visualização para: {os.path.basename(file_path)}" + (f" - Aba: {sheet_name}" if sheet_name else ""), LogLevel.INFO)
        try:
            n_preread_rows = 20
            header_row_index = 0
            df_preview_sliced = None
            
            # 1. Pré-leitura para detectar o cabeçalho
            pre_read_df = None
            delimiter = self.get_selected_delimiter() if file_path.lower().endswith((".csv", ".txt")) else None
            
            if file_path.lower().endswith((".csv", ".txt")):
                if not delimiter:
                    self.log_message("Pré-visualização falhou: Delimitador inválido.", LogLevel.ERROR)
                    return
                pre_read_df = pl.read_csv(source=file_path, has_header=False, n_rows=n_preread_rows, separator=delimiter, encoding='latin-1', ignore_errors=True, infer_schema = False, quote_char = None, truncate_ragged_lines = True)
            elif file_path.lower().endswith((".xlsx", ".xls")) and sheet_name:
                pre_read_df = pl.read_excel(source=file_path, sheet_name=sheet_name, has_header = False).head(n_preread_rows)

            if pre_read_df is not None and not pre_read_df.is_empty():
                header_row_index = _find_header_row_index(pre_read_df, n_preread_rows)

                # 2. Extrair cabeçalhos, dados e renomear (a lógica robusta)
                header_names_raw = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(pre_read_df.row(header_row_index))]
                header_names = _make_headers_unique(header_names_raw)
                data_rows = pre_read_df.slice(offset=header_row_index + 1).head(n_rows_to_preview)
                
                if not data_rows.is_empty():
                    rename_mapping = {old_name: new_name for old_name, new_name in zip(data_rows.columns, header_names)}
                    df_preview_sliced = data_rows.rename(rename_mapping)

            # 3. Carregar os dados no modelo da tabela
            if df_preview_sliced is not None and not df_preview_sliced.is_empty():
                self.preview_table_model.load_data(df_preview_sliced)
                self.log_message(f"Pré-visualização gerada com {df_preview_sliced.height} linhas.", LogLevel.SUCCESS)
            else:
                self.log_message(f"O arquivo/aba {os.path.basename(file_path)} está vazio ou não foi possível ler dados para pré-visualização.", LogLevel.WARNING)
                self.preview_table_model.clear_data()

        except Exception as e:
            self.log_message(f"Erro ao gerar pré-visualização para {os.path.basename(file_path)}: {e}", LogLevel.ERROR)
            self.preview_table_model.clear_data()
    
    def open_filter_dialog(self):
        if not self.header_mapping:
            self.log_message("Por favor, analise e mapeie os cabeçalhos primeiro.", LogLevel.WARNING)
            return
        final_headers = {
            map_info['final_name'] for map_info in self.header_mapping.values() if map_info.get('include')
        }
        if not final_headers:
            self.log_message("Nenhym cabeçalho final encontrado no mapeamento. Impossível definir filtros", LogLevel.WARNING)
            return
        dialog = FilterDialog(list(final_headers), self.filter_rules, self)
        if dialog.exec() == QDialog.Accepted:
            self.filter_rules = dialog.get_filters()
            self.log_message(f"Filtros atualizados. {len(self.filter_rules)} regra(s) ativa(s)", LogLevel.SUCCESS)
        else:
            self.log_message("Definição de filtros cancelada.", LogLevel.INFO)

    def on_file_selected(self, current_file_item, previous_file_item):
        """Chamado quando um arquivo é selecionado na lista de arquivos. Delega o carregamento de abas para uma thread."""
        self.sheets_list_widget.clear()
        self.sheets_list_widget.setEnabled(False)
        self.mark_all_sheets_button.setEnabled(False)
        self.unmark_all_sheets_button.setEnabled(False)
        # self.preview_table_model.clear_data() # A pré-visualização é tratada por on_file_selected_for_preview

        if self.sheet_loader_thread and self.sheet_loader_thread.isRunning():
            self.sheet_loader_thread.stop() # Solicita parada da thread anterior
            # Não esperar aqui para não bloquear a UI, a thread antiga morrerá ou emitirá um resultado que será ignorado

        if not current_file_item:
            return

        selected_file_name = current_file_item.text()
        file_path = self.current_files_paths.get(selected_file_name)

        if not file_path:
            self.log_message(f"Caminho não encontrado para o arquivo: {selected_file_name}", LogLevel.ERROR)
            return

        # CSVs não têm abas, então não iniciar a thread para eles aqui.
        # A lógica de pré-visualização de CSV é tratada em on_file_selected_for_preview
        if file_path.lower().endswith((".xlsx", ".xls")):
            self.log_message(f"Carregando abas para '{selected_file_name}'...", LogLevel.INFO)
            self.sheets_list_widget.setEnabled(False) # Garante que esteja desabilitada
            self.mark_all_sheets_button.setEnabled(False)
            self.unmark_all_sheets_button.setEnabled(False)

            self.sheet_loader_thread = SheetLoadingWorker(file_path)
            self.sheet_loader_thread.finished.connect(self.on_sheet_loading_finished)
            self.sheet_loader_thread.start()
        else: # Arquivo não-Excel (ex: CSV) selecionado
            self.log_message(f"Arquivo selecionado: {selected_file_name} (Não é Excel, sem abas para listar).", LogLevel.INFO)
            # A pré-visualização de CSV será acionada por on_file_selected_for_preview
            # Nenhuma aba para popular, então botões de aba permanecem desabilitados.


    def on_sheet_loading_finished(self, file_path_processed, sheet_names, error_message):
        """Chamado quando a SheetLoadingWorker termina de carregar as abas."""
        # Verificar se o resultado ainda é para o arquivo atualmente selecionado
        current_selected_file_item = self.files_list_widget.currentItem()
        if not current_selected_file_item or self.current_files_paths.get(current_selected_file_item.text()) != file_path_processed:
            self.log_message(f"Resultado do carregamento de abas para '{os.path.basename(file_path_processed)}' ignorado (seleção mudou).", LogLevel.INFO)
            if self.sheet_loader_thread and self.sheet_loader_thread.file_path == file_path_processed:
                self.sheet_loader_thread = None # Limpa a referência da thread que acabou
            return

        selected_file_name = os.path.basename(file_path_processed) # Obter nome do arquivo do caminho processado

        if error_message:
            self.log_message(error_message, LogLevel.ERROR)
            if file_path_processed in self.sheet_selections: # Limpar se deu erro
                del self.sheet_selections[file_path_processed]
            self.sheets_list_widget.clear() # Garante que a lista de abas esteja limpa
            self.sheets_list_widget.setEnabled(False)
            self.mark_all_sheets_button.setEnabled(False)
            self.unmark_all_sheets_button.setEnabled(False)
            if self.sheet_loader_thread and self.sheet_loader_thread.file_path == file_path_processed:
                self.sheet_loader_thread = None
            return

        # Sucesso no carregamento das abas
        self.log_message(f"Abas carregadas para '{selected_file_name}'.", LogLevel.INFO)
        can_manage_sheets = False
        if sheet_names:
            self.sheets_list_widget.setEnabled(True)
            can_manage_sheets = True

            # Desconectar temporariamente o sinal para evitar chamadas recursivas
            try:
                self.sheets_list_widget.itemChanged.disconnect(self.on_sheet_selection_changed)
            except RuntimeError:
                pass

            # Verificar se já temos seleções salvas para este arquivo
            if file_path_processed not in self.sheet_selections:
                self.sheet_selections[file_path_processed] = {name: True for name in sheet_names}

            current_sheet_states = self.sheet_selections[file_path_processed]

            # Adicionar abas que estão no arquivo, mas talvez não em current_sheet_states (se o arquivo mudou)
            for sheet_name in sheet_names:
                if sheet_name not in current_sheet_states:
                    current_sheet_states[sheet_name] = True # Default para incluir novas abas

            # Remover abas de current_sheet_states que não existem mais no arquivo
            # (Isso garante que self.sheet_selections reflita apenas abas existentes)
            # Criar uma cópia das chaves para iterar, pois podemos modificar o dicionário
            for stored_sheet_name in list(current_sheet_states.keys()):
                if stored_sheet_name not in sheet_names:
                    del current_sheet_states[stored_sheet_name]

            self.sheets_list_widget.clear() # Limpar antes de repopular
            for sheet_name in sheet_names: # Iterar sobre as abas reais do arquivo
                item = QListWidgetItem(sheet_name)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                is_checked = current_sheet_states.get(sheet_name, True) # Default para True
                item.setCheckState(Qt.Checked if is_checked else Qt.Unchecked)
                self.sheets_list_widget.addItem(item)

            log_suffix = " (carregadas do cache/atualizadas)" if any(not s for s in current_sheet_states.values()) else " (todas marcadas por padrão/atualizadas)"
            self.log_message(f"Abas encontradas para '{selected_file_name}': {', '.join(sheet_names)}.{log_suffix}", LogLevel.INFO)

            # Reconectar o sinal
            self.sheets_list_widget.itemChanged.connect(self.on_sheet_selection_changed)

            # Tenta pré-selecionar e pré-visualizar a primeira aba, se houver
            if self.sheets_list_widget.count() > 0:
                first_sheet_item = self.sheets_list_widget.item(0)
                if first_sheet_item:
                    # Se on_file_selected_for_preview também chama on_file_selected,
                    # precisamos garantir que on_file_selected não reinicie a thread de abas desnecessariamente.
                    # No entanto, on_file_selected_for_preview chama on_file_selected PRIMEIRO.
                    # A chamada a setCurrentItem aqui irá disparar on_sheet_list_item_selected_for_preview, que é para a pré-visualização.
                    self.sheets_list_widget.setCurrentItem(first_sheet_item)


        else: # Nenhuma aba encontrada
            self.log_message(f"Nenhuma aba encontrada no arquivo Excel: {selected_file_name}", LogLevel.WARNING)
            if file_path_processed in self.sheet_selections:
                del self.sheet_selections[file_path_processed]
            self.sheets_list_widget.clear()
            self.sheets_list_widget.setEnabled(False)

        self.mark_all_sheets_button.setEnabled(can_manage_sheets)
        self.unmark_all_sheets_button.setEnabled(can_manage_sheets)

        if self.sheet_loader_thread and self.sheet_loader_thread.file_path == file_path_processed:
            self.sheet_loader_thread = None # Limpa a referência da thread que acabou

    def on_sheet_selection_changed(self, item_changed: QListWidgetItem):
        """Chamado quando o estado de um checkbox de aba muda."""
        current_file_item = self.files_list_widget.currentItem()
        if not current_file_item:
            return # Nenhum arquivo Excel selecionado na lista principal

        selected_file_name = current_file_item.text()
        file_path = self.current_files_paths.get(selected_file_name)

        if not file_path or not file_path.lower().endswith((".xlsx", ".xls")):
            return # Não é um arquivo Excel válido ou caminho não encontrado

        sheet_name = item_changed.text()
        is_checked = item_changed.checkState() == Qt.Checked

        # Garantir que a entrada para o arquivo exista em self.sheet_selections
        if file_path not in self.sheet_selections:
            # Isso não deveria acontecer se on_file_selected populou corretamente,
            # mas é uma salvaguarda.
            self.sheet_selections[file_path] = {} 
        
        self.sheet_selections[file_path][sheet_name] = is_checked
        # self.log_message(f"Seleção da aba '{sheet_name}' para '{selected_file_name}' atualizada para: {'Marcada' if is_checked else 'Desmarcada'}", LogLevel.INFO)

    def get_files_and_sheets_to_process(self):
        files_to_process_list = []
        if not self.current_files_paths:
            self.log_message("Nenhuma pasta selecionada ou nenhum arquivo encontrado.", LogLevel.WARNING)
            return None

        for file_name, file_path in self.current_files_paths.items():
            is_excel = file_path.lower().endswith((".xlsx", ".xls"))

            if is_excel:
                selected_sheets_for_file = []
                
                # --- NOVA LÓGICA DE DECISÃO ---
                # Se existem regras globais, use-as.
                if self.sheet_selection_rules and self.all_sheets_cache:
                    file_sheets = self.all_sheets_cache.get(file_path, [])
                    rule_mode = self.sheet_selection_rules.get("mode", "include")
                    rule_names = self.sheet_selection_rules.get("names", set())

                    if rule_mode == "include":
                        selected_sheets_for_file = [s for s in file_sheets if s in rule_names]
                    elif rule_mode == "exclude":
                        selected_sheets_for_file = [s for s in file_sheets if s not in rule_names]
                
                # Senão, use a lógica antiga de seleção manual por arquivo.
                else:
                    if file_path in self.sheet_selections:
                        selected_sheets_for_file = [s for s, checked in self.sheet_selections[file_path].items() if checked]
                    # Se um arquivo nunca foi selecionado, pode não ter entrada. Neste caso, não processará nenhuma aba.
                    # Isso é melhor do que o comportamento antigo que tentava ler todas.
                
                if not selected_sheets_for_file:
                    self.log_message(f"Nenhuma aba selecionada para o arquivo Excel '{file_name}'. Será pulado.", LogLevel.INFO)
                    continue
                
                files_to_process_list.append((file_path, selected_sheets_for_file))

            elif file_path.lower().endswith((".csv", ".txt")):
                files_to_process_list.append((file_path, None))
            else:
                self.log_message(f"Arquivo '{file_name}' não é suportado. Pulando.", LogLevel.WARNING)

        if not files_to_process_list:
            self.log_message("Nenhum arquivo encontrado ou selecionado para processamento após aplicar as regras.", LogLevel.WARNING)
            return None
            
        return files_to_process_list


    def log_message(self, message, level=LogLevel.INFO):
        self.is_last_log_progress = False
        self.log_console_text_edit.append(f"{level.value} {message}")

    def list_files_in_folder(self, folder_path):
        self.files_list_widget.clear()
        self.sheet_selections.clear() 
        self.sheets_list_widget.clear()
        self.sheets_list_widget.setEnabled(False)
        self.current_files_paths.clear()
        self.preview_table_model.clear_data()

        self.refresh_button.setEnabled(False) 
        self.map_headers_button.setEnabled(False)
        self.define_filters_button.setEnabled(False) 
        self.sheet_selection_button.setEnabled(False)
        self.pivot_button.setEnabled(False)
        self.header_mapping.clear()
        self.filter_rules.clear()
        self.pivot_rules.clear()
        self.duplicate_key_columns.clear()
        self.sheet_selection_rules.clear()
        self.all_sheets_cache.clear()


        supported_extensions = ("*.xlsx", "*.csv", "*.xls", "*.txt")
        found_files_paths = []
        try:
            for ext in supported_extensions:
                pattern = os.path.join(folder_path, ext)
                found_files_paths.extend(glob.glob(pattern))
            
            if found_files_paths:
                for full_path in found_files_paths:
                    file_name = os.path.basename(full_path)
                    self.files_list_widget.addItem(file_name)
                    self.current_files_paths[file_name] = full_path
                self.log_message(f"Encontrados {len(found_files_paths)} arquivos na pasta.", LogLevel.SUCCESS)
                self.map_headers_button.setEnabled(True) # HABILITAR AQUI se arquivos forem encontrados
                self.refresh_button.setEnabled(True) # E então ele é habilitado
                excel_files_found = any(f.lower().endswith(('.xlsx', '.xls')) for f in found_files_paths)
                self.sheet_selection_button.setEnabled(excel_files_found)
            else:
                self.log_message("Nenhum arquivo suportado (.xlsx, .csv, .xls, .txt) encontrado na pasta.", LogLevel.WARNING)
                # self.map_headers_button.setEnabled(False) # Já está desabilitado pelo início da função

        except Exception as e:
            self.log_message(f"Erro ao listar arquivos: {e}", LogLevel.ERROR)
            self.current_files_paths.clear()
            # self.map_headers_button.setEnabled(False) # Já está desabilitado
    
    def update_output_filename_extension(self, selected_format):
        current_name = self.output_name_line_edit.text()
        name_part, _ = os.path.splitext(current_name)
        new_extension = ".parquet" if selected_format == "Parquet" else "." + selected_format.lower()
        # new_extension = "." + selected_format.lower()
        self.output_name_line_edit.setText(name_part + new_extension)

    def open_save_file_dialog(self):
        current_folder = self.folder_path_line_edit.text() or os.path.expanduser("~")
        suggested_filename = self.output_name_line_edit.text()
        initial_path = os.path.join(current_folder, suggested_filename)
        selected_format = self.output_format_combo_box.currentText()
        if selected_format == "XLSX":
            filter_str = "Arquivos Excel (*.xlsx)"
        elif selected_format == "CSV":
            filter_str = "Arquivos CSV (*.csv)"
        elif selected_format == "Parquet":
            filter_str = "Arquivos Parquet (*.parquet)"
        else:
            filter_str = "Todos os Arquivos (*)"
        file_path, _ = QFileDialog.getSaveFileName(self, "Salvar Arquivo Consolidado Como...", initial_path, filter_str)
        if file_path:
            self.output_file_path = file_path
            base_name = os.path.basename(file_path)
            self.output_name_line_edit.setText(base_name)
            name_part, ext_part = os.path.splitext(base_name)
            if ext_part.lower() == ".xlsx" and self.output_format_combo_box.currentText() != "XLSX":
                self.output_format_combo_box.setCurrentText("XLSX")
            elif ext_part.lower() == ".csv" and self.output_format_combo_box.currentText() != "CSV":
                self.output_format_combo_box.setCurrentText("CSV")
            elif ext_part.lower() == ".parquet" and self.output_format_combo_box.currentText != "Parquet":
                self.output_format_combo_box.setCurrentText("Parquet")
            self.log_message(f"Arquivo de saída definido como: {file_path}", LogLevel.SUCCESS)
        else:
            self.log_message("Seleção de local para salvar cancelada.", LogLevel.INFO)
    
    def update_progress_text(self, text):
        """Atualiza a última linha do console com o texto de progresso, ou adiciona uma nova linha se a última não
        for de progresso."""
        formatted_text = f"{LogLevel.INFO.value} {text}"
        cursor = self.log_console_text_edit.textCursor()
        if self.is_last_log_progress:
            cursor.movePosition(QTextCursor.End)
            cursor.select(QTextCursor.BlockUnderCursor)
            cursor.removeSelectedText()
        # cursor.insertText(formatted_text)
        cursor.movePosition(QTextCursor.End)
        self.log_console_text_edit.append(formatted_text)
        self.log_console_text_edit.ensureCursorVisible()
        self.is_last_log_progress = True

    def start_consolidation(self):
        if not self.folder_path_line_edit.text():
            self.log_message("Por favor, selecione uma pasta de projeto primeiro.", LogLevel.WARNING)
            return

        if not self.output_file_path:
            self.log_message("Por favor, defina o arquivo de saída usando 'Salvar Como...'.", LogLevel.WARNING)
            return

        files_to_process = self.get_files_and_sheets_to_process()
        if not files_to_process: # Checa se é None ou lista vazia
            self.log_message("Nenhum arquivo ou aba válida selecionada para consolidação.", LogLevel.WARNING)
            return

        self.set_ui_for_processing(True)
        output_format = self.output_format_combo_box.currentText()
        selected_delimiter = self.get_selected_delimiter()
        if not selected_delimiter:
            self.log_message("Delimitador inválido ou não definido para arquivos CSV/TXT.", LogLevel.ERROR)
            self.set_ui_for_processing(False)
            return
        
        
        self.consolidation_thread = ConsolidationWorker(files_to_process, self.output_file_path, output_format, self.header_mapping, self.filter_rules, selected_delimiter, self.pivot_rules, self.duplicates_config)
        self.consolidation_thread.log_message.connect(self.log_message) 
        self.consolidation_thread.progress_updated.connect(self.update_progress_bar)
        self.consolidation_thread.finished.connect(self.on_consolidation_finished)
        self.consolidation_thread.progress_text_updated.connect(self.update_progress_text)
        
        self.consolidation_thread.start()

    def update_progress_bar(self, value):
        self.progress_bar.setValue(value)

    def on_consolidation_finished(self, success, message):
        if success:
            self.log_message(f"Resultado: {message}", LogLevel.SUCCESS)
        else:
            self.log_message(f"Resultado: {message}", LogLevel.ERROR)
        
        self.set_ui_for_processing(False)
        self.consolidation_thread = None 

    def set_ui_for_processing(self, processing):
        not_proc = not processing
        self.select_folder_button.setEnabled(not_proc)
        self.files_list_widget.setEnabled(not_proc)
        
        # Habilitar o botão de mapear apenas se não estiver processando E houver arquivos listados
        self.map_headers_button.setEnabled(not_proc and self.files_list_widget.count() > 0) 

        has_sel_excel_sheets = False
        current_file_item = self.files_list_widget.currentItem()
        if not_proc and current_file_item: # Só verifica se não estiver processando e houver item
            fn = current_file_item.text()
            fp = self.current_files_paths.get(fn)
            if fp and (fp.lower().endswith((".xlsx",".xls"))) and self.sheets_list_widget.count() > 0:
                has_sel_excel_sheets = True
        self.sheets_list_widget.setEnabled(not_proc and has_sel_excel_sheets)
        sheet_buttons_enabled = not_proc and self.sheets_list_widget.isEnabled() and self.sheets_list_widget.count() > 0
        self.mark_all_sheets_button.setEnabled(sheet_buttons_enabled)
        self.unmark_all_sheets_button.setEnabled(sheet_buttons_enabled)
        self.output_name_line_edit.setEnabled(not_proc)
        self.output_format_combo_box.setEnabled(not_proc)
        self.save_as_button.setEnabled(not_proc)
        self.consolidate_button.setVisible(not_proc) 
        self.cancel_button.setVisible(processing)
        # self.progress_bar.setVisible(processing)
        # self.progress_text_label.setVisible(processing)       
        if not_proc:
            self.progress_bar.setValue(0)
        self.is_last_log_progress = False

    def cancel_consolidation(self):
        if self.consolidation_thread and self.consolidation_thread.isRunning():
            self.log_message("Solicitando cancelamento da consolidação...", LogLevel.INFO)
            self.consolidation_thread.stop() 
        else:
            self.log_message("Nenhuma consolidação em andamento para cancelar.", LogLevel.INFO)
            self.set_ui_for_processing(False) # Garante que a UI volte ao normal se o botão for clicado por engano

    def open_folder_dialog(self): # Adicionado para garantir que está presente
        start_dir = self.folder_path_line_edit.text() or os.path.expanduser("~")
        folder_path = QFileDialog.getExistingDirectory(self, "Selecionar Pasta do Projeto", start_dir)
        if folder_path:
            self.folder_path_line_edit.setText(folder_path)
            self.log_message(f"Pasta selecionada: {folder_path}", LogLevel.INFO)
            self.list_files_in_folder(folder_path)
        else:
            self.log_message("Seleção de pasta cancelada.", LogLevel.INFO)

    def closeEvent(self, event):
        if self.consolidation_thread and self.consolidation_thread.isRunning():
            self.log_message("Fechando aplicação, parando consolidação em andamento...", LogLevel.INFO)
            self.consolidation_thread.stop()
            self.consolidation_thread.wait() 

        if self.sheet_loader_thread and self.sheet_loader_thread.isRunning(): # Adicionado
            self.log_message("Fechando aplicação, parando carregamento de abas...", LogLevel.INFO)
            self.sheet_loader_thread.stop()
            self.sheet_loader_thread.wait() # Espera a thread finalizar
        
        if self.header_analyzer_thread and self.header_analyzer_thread.isRunning(): # Adicionado
            self.log_message("Fechando aplicação, parando análise de cabeçalhos...", LogLevel.INFO)
            self.header_analyzer_thread.stop()
            self.header_analyzer_thread.wait()

        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    app.setStyle("Fusion")

    # Paleta de cores para o tema claro e profissional
    light_palette = QPalette()
    
    # Cores de fundo e texto gerais
    light_palette.setColor(QPalette.Window, QColor("#F0F0F0")) # Fundo principal cinza bem claro
    light_palette.setColor(QPalette.WindowText, QColor("#333333")) # Texto escuro (não preto)
    
    # Cores para campos de entrada, listas, tabelas
    light_palette.setColor(QPalette.Base, QColor("#FFFFFF")) # Fundo branco para inputs
    light_palette.setColor(QPalette.AlternateBase, QColor("#F5F5F5")) # Cor de linha alternada
    light_palette.setColor(QPalette.Text, QColor("#333333")) # Texto dentro dos inputs
    
    # Cores de botões
    light_palette.setColor(QPalette.Button, QColor("#E0E0E0")) # Botões cinza claro
    light_palette.setColor(QPalette.ButtonText, QColor("#333333")) # Texto dos botões
    
    # Cores de destaque (seleção, links, etc.)
    CORPORATE_GREEN = QColor("#86BC25")
    light_palette.setColor(QPalette.Highlight, CORPORATE_GREEN)
    light_palette.setColor(QPalette.HighlightedText, Qt.white)
    
    # Aplicar a paleta de cores base
    app.setPalette(light_palette)

    # Folha de estilos QSS para refinamentos
    professional_stylesheet = """
        QWidget {
            font-family: Aptos, Segoe UI, Arial, sans-serif; /* Define a fonte com fallbacks */
            font-size: 9pt;
        }
        QGroupBox {
            font-weight: bold;
            font-size: 10pt;
        }
        QPushButton {
            border: 1px solid #B0B0B0;
            border-radius: 4px;
            padding: 6px;
            background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                              stop: 0 #F5F5F5, stop: 1 #E0E0E0);
            min-width: 80px;
        }
        QPushButton:hover {
            background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                              stop: 0 #FFFFFF, stop: 1 #E8E8E8);
        }
        QPushButton:pressed {
            background-color: #D0D0D0;
        }
        QPushButton:disabled {
            background-color: #E0E0E0;
            color: #A0A0A0;
            border-color: #C0C0C0;
        }
        /* Estilo especial para o botão de consolidação */
        #consolidate_button {
            font-weight: bold;
            color: white;
            background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                              stop: 0 #86BC25, stop: 1 #70A01C);
            border-color: #608918;
        }
        #consolidate_button:hover {
            background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                              stop: 0 #92CC2A, stop: 1 #7CB820);
        }
        #consolidate_button:pressed {
            background-color: #70A01C;
        }

        QLineEdit, QListWidget, QTableWidget, QTextEdit {
            border: 1px solid #B0B0B0;
            border-radius: 4px;
            padding: 4px;
            background-color: #FFFFFF;
        }
        QListWidget {
            border: 1px solid #B0B0B0;
            border-radius: 4px;
            padding: 2px;
            background-color: #FFFFFF;
        }
        QListWidget::item:hover {
            background-color: #F0F0F0;
            border-radius: 3px;
        }
        QComboBox {
            border: 1px solid #B0B0B0;
            border-radius: 4px;
            padding: 4px;
            padding-right: 25px;
            background-color: #FFFFFF;
        }
        QComboBox::drop-down {
            subcontrol-origin: padding;
            subcontrol-position: top right;
            width: 25px;
            border-left-width: 1px;
            border-left-color: #B0B0B0;
            border-left-style: solid;
            border-top-right-radius: 3px;
            border-bottom-right-radius: 3px;
        }
        QComboBox::down-arrow{
            image: url(app/down.png);
            width: 14px;
            height: 14px;
        }
        QComboBox::drop-down:hover{
            background-color: #E8E8E8
        }
        QComboBox, QAbstractItemView::item:selected {
            background-color: #FFFFFF;
            color: black;
        }
        QHeaderView::section {
            background-color: #4A5568; /* Cinza-azulado escuro corporativo */
            color: white;
            padding: 4px;
            border: 1px solid #E0E0E0;
            font-weight: bold;
        }
        QProgressBar {
            border: 1px solid #B0B0B0;
            border-radius: 5px;
            text-align: center;
            color: #333333;
        }
        QProgressBar::chunk {
            background-color: #86BC25; /* Verde corporativo */
            border-radius: 4px;
        }
    """
    app.setStyleSheet(professional_stylesheet)
    
    window = MainWindow()
    window.consolidate_button.setObjectName("consolidate_button")
    sys.exit(app.exec())