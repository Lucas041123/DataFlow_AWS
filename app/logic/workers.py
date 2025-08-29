import os
import polars as pl
import openpyxl
import xlrd
from collections import defaultdict, Counter

from PySide6.QtCore import QThread, Signal

# Importa as funções e constantes do novo módulo de utilitários
from ..utils import (
    LogLevel, _find_header_row_index, _make_headers_unique, 
    _normalize_header_name, DATA_TYPES_OPTIONS, TYPE_STRING_TO_POLARS,
    OPERATOR_OPTIONS, OPERATORS_NO_VALUE
)

# Exceção customizada para interrupção
class InterruptedError(Exception):
    pass

class ConsolidationWorker(QThread):
    progress_updated = Signal(int) 
    log_message = Signal(str, LogLevel) 
    finished = Signal(bool, str) 
    progress_text_updated = Signal(str)

    def __init__(self, files_to_process, output_path, output_format, header_mapping, filter_rules, delimiter, pivot_rules, duplicates_config=None):
        super().__init__()
        self.files_to_process = files_to_process 
        self.output_path = output_path
        self.output_format = output_format
        self.header_mapping = header_mapping
        self.filter_rules = filter_rules
        self.pivot_rules = pivot_rules
        self.duplicates_config = duplicates_config or {}
        self.delimiter = delimiter
        self.is_running = True

    def run(self):
        try:
            self.log_message.emit("Iniciando processo de consolidação...", LogLevel.INFO)
            all_dataframes_processed = [] 
            
            total_items = 0
            for _, sheets_to_process_for_file in self.files_to_process:
                if sheets_to_process_for_file is None: total_items += 1
                else: total_items += len(sheets_to_process_for_file)
            if total_items == 0:
                self.log_message.emit("Nenhum item válido para processar.", LogLevel.WARNING)
                self.finished.emit(False, "Nenhum item para processar.")
                return
            processed_items = 0

            for file_path, selected_sheets in self.files_to_process:
                if not self.is_running: break 
                file_name = os.path.basename(file_path)
                sheets_to_iterate = selected_sheets if selected_sheets is not None else [None]

                for sheet_name in sheets_to_iterate:
                    if not self.is_running: break 
                    current_item_description = f"'{file_name}'" + (f" - Aba: '{sheet_name}'" if sheet_name else "")
                    try:
                        # ETAPA 1: Pré-leitura e Detecção do Cabeçalho
                        n_preread_rows = 20
                        header_row_index = 0
                        header_names = []
                        
                        pre_read_df = None
                        if file_path.lower().endswith((".csv", ".txt")):
                            pre_read_df = pl.read_csv(source=file_path, has_header=False, n_rows=n_preread_rows, separator=self.delimiter, encoding='latin-1', ignore_errors=True, infer_schema = False, quote_char = None, truncate_ragged_lines = True)
                        elif file_path.lower().endswith((".xlsx", ".xls")):
                            pre_read_df = pl.read_excel(source=file_path, sheet_name=sheet_name, has_header = False).head(n_preread_rows)

                        if pre_read_df is not None and not pre_read_df.is_empty():
                            header_row_index = _find_header_row_index(pre_read_df, n_preread_rows)
                            header_names_raw = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(pre_read_df.row(header_row_index))]
                            header_names = _make_headers_unique(header_names_raw)
                        
                        # --- LÓGICA DE LEITURA FINAL E ROBUSTA ---
                        df_original = None
                        
                        # Ler o arquivo inteiro como dados brutos, sem cabeçalho
                        if file_path.lower().endswith((".csv", ".txt")):
                            df_raw_data = pl.read_csv(source=file_path, has_header=False, separator=self.delimiter, encoding='latin-1', infer_schema = False, ignore_errors = True, quote_char = None, truncate_ragged_lines = True)
                        elif file_path.lower().endswith((".xlsx", ".xls")):
                            df_raw_data = pl.read_excel(source=file_path, sheet_name=sheet_name, has_header = False)
                        
                        if df_raw_data is not None and not df_raw_data.is_empty():
                            # Fatiar o DataFrame para remover lixo + linha do cabeçalho
                            df_data_only = df_raw_data.slice(offset=header_row_index + 1)
                            
                            if not df_data_only.is_empty():
                                # Renomear as colunas com os nomes que detectamos
                                rename_mapping = {old_name: new_name for old_name, new_name in zip(df_data_only.columns, header_names)}
                                df_original = df_data_only.rename(rename_mapping)

                        # Sua lógica de fallback para erros de conversão no Excel pode ser integrada aqui se necessário,
                        # mas esta leitura primária é muito mais estável.
                        
                        if df_original is None or df_original.is_empty():
                            self.log_message.emit(f"Dados vazios ou erro ao ler {current_item_description}. Pulando.", LogLevel.WARNING)
                            processed_items += 1
                            continue

                        # --- 1. Aplicar Mapeamento de Nomes e Filtro de Colunas (com Coalesce) ---
                        df_intermediate = df_original
                        if self.header_mapping:
                            # 1. Agrupar colunas de origem por seu nome final de destino
                            final_name_to_source = defaultdict(list)
                            for original_col_name in df_original.columns:
                                source_key = (original_col_name, file_path, sheet_name)
                                mapping_info = self.header_mapping.get(source_key)
                                if mapping_info and mapping_info.get("include", False):
                                    final_name = mapping_info.get("final_name")
                                    final_name_to_source[final_name].append(original_col_name)
                            # final_name_to_original_map = {}
                            # for original_name, mapping_info in self.header_mapping.items():
                            #     if mapping_info.get("include", False):
                            #         final_name = mapping_info.get("final_name", original_name)
                            #         if final_name not in final_name_to_original_map:
                            #             final_name_to_original_map[final_name] = []
                            #         final_name_to_original_map[final_name].append(original_name)

                            # 2. Construir as expressões de seleção usando coalesce quando necessário
                            select_expressions = []
                            df_original_columns = set(df_original.columns) # Para buscas mais rápidas

                            for final_name, original_cols_list in final_name_to_source.items():
                                # Filtrar a lista para incluir apenas as colunas que existem no DataFrame atual
                                # existing_original_cols = [col for col in original_cols_list if col in df_original_columns]
                                
                                # if not existing_original_cols:
                                #     continue # Nenhuma das colunas de origem para este nome final existe neste arquivo/aba

                                # Se apenas uma coluna de origem existe, faz um alias simples.
                                # Se mais de uma, usa coalesce para combinar os dados.
                                if not original_cols_list:
                                    continue
                                if len(original_cols_list) > 1:
                                    self.log_message.emit(f"Combinando colunas {original_cols_list} em '{final_name}' para {current_item_description}", LogLevel.INFO)
                                    expr = pl.coalesce(original_cols_list).alias(final_name)
                                else:
                                    self.log_message.emit(f"Combinando colunas {original_cols_list} em '{final_name}' para {current_item_description}", LogLevel.INFO)
                                    expr = pl.col(original_cols_list[0]).alias(final_name)
                                
                                select_expressions.append(expr)

                            # Se, após o mapeamento, não sobrar nenhuma expressão, pular o arquivo/aba
                            if not select_expressions:
                                self.log_message.emit(f"Nenhuma coluna do arquivo {current_item_description} corresponde ao mapeamento. Pulando.", LogLevel.WARNING)
                                processed_items += 1
                                continue

                            # 3. Executar a seleção no DataFrame
                            df_intermediate = df_original.select(select_expressions)
                        
                        if df_intermediate.width == 0:
                            self.log_message.emit(f"Nenhuma coluna restante em {current_item_description} após mapeamento de nomes. Pulando.", LogLevel.WARNING)
                            processed_items += 1; continue
                        
                        # --- 2. Aplicar Tipagem Especificada pelo Usuário ---
                        df_typed = df_intermediate
                        if self.header_mapping:
                            # Precisamos iterar sobre as colunas FINAIS do df_intermediate
                            # e encontrar a regra de tipagem correspondente no header_mapping
                            # (que é chaveado pelo nome ORIGINAL).
                            
                            # Criar um mapa reverso: {final_name: original_name_info_from_mapping}
                            # ou mais direto: {final_name: type_str_from_mapping}
                            final_name_to_type_str = {}
                            for original_h, map_details in self.header_mapping.items():
                                if map_details.get("include"):
                                    final_name_to_type_str[map_details.get("final_name", original_h)] = map_details.get("type_str")

                            casting_expressions = []
                            for final_col_name in df_typed.columns: # Iterar sobre colunas já mapeadas/renomeadas
                                type_str = final_name_to_type_str.get(final_col_name)
                                
                                if type_str and type_str != DATA_TYPES_OPTIONS[0]: # Se não for "Automático/String"
                                    polars_type = TYPE_STRING_TO_POLARS.get(type_str)
                                    if polars_type:
                                        self.log_message.emit(f"Convertendo coluna '{final_col_name}' para {type_str} em {current_item_description}", LogLevel.INFO)
                                        casting_expressions.append(pl.col(final_col_name).cast(polars_type, strict=False))
                                    else: # Tipo string não encontrado no mapa (não deve acontecer)
                                        casting_expressions.append(pl.col(final_col_name)) # Manter como está
                                else: # "Automático/String" ou tipo não mapeado
                                    casting_expressions.append(pl.col(final_col_name)) # Manter como está
                            
                            if casting_expressions: # Se houver alguma expressão (sempre haverá se df_typed não for vazio)
                                df_typed = df_typed.select(casting_expressions) # .select() é mais seguro que .with_columns() para recriar
                    
                        # --- 4. Aplicar Filtros (com lógica hierárquica E/OU) ---
                        df_filtered = df_typed
                        if self.filter_rules:
                            # from collections import defaultdict

                            # Definir quais operadores são para exclusão
                            EXCLUSION_OPERATORS = {"Diferente de", "Não contém"}

                            grouped_rules = defaultdict(list)
                            for rule in self.filter_rules:
                                if rule.get("column"):
                                    grouped_rules[rule["column"]].append(rule)

                            final_expressions_to_and = []
                            df_schema = df_filtered.schema

                            for col_name, rules_for_col in grouped_rules.items():
                                if col_name not in df_schema:
                                    continue

                                inclusion_exprs = []
                                exclusion_exprs = []
                                col_type = df_schema[col_name]

                                # 1. Separar regras em Inclusão e Exclusão
                                for rule in rules_for_col:
                                    operator = rule.get("operator")
                                    value = rule.get("value") # Pega o valor (pode ser string ou lista)

                                    target_list = exclusion_exprs if operator in EXCLUSION_OPERATORS else inclusion_exprs

                                    # Adicionado para pular regras incompletas
                                    if operator is None or value is None:
                                        continue
                                    
                                    try:
                                        polars_col = pl.col(col_name)
                                        expr = None
                                        
                                        if operator in OPERATORS_NO_VALUE:
                                            if operator == "Está em branco": expr = polars_col.is_null()
                                            elif operator == "Não está em branco": expr = polars_col.is_not_null()
                                        
                                        elif operator == "Entre":
                                            if isinstance(value, list) and len(value) == 2:
                                                min_val_str, max_val_str = value
                                                # Strip é aplicado aqui, onde sabemos que são strings
                                                if min_val_str.strip() and max_val_str.strip():
                                                    lit_min = pl.lit(min_val_str.strip()).cast(col_type, strict=False)
                                                    lit_max = pl.lit(max_val_str.strip()).cast(col_type, strict=False)
                                                    expr = polars_col.is_between(lit_min, lit_max)

                                        # Garante que o valor é uma string antes de usar o .strip()
                                        elif isinstance(value, str) and value.strip():
                                            value_str = value.strip()
                                            lit_val = pl.lit(value_str).cast(col_type, strict=False)

                                            if operator == "Igual a": expr = (polars_col == lit_val)
                                            elif operator == "Diferente de": expr = (polars_col != lit_val)
                                            elif operator == "Maior que": expr = (polars_col > lit_val)
                                            elif operator == "Menor que": expr = (polars_col < lit_val)
                                            elif col_type == pl.String:
                                                if operator == "Contém": expr = polars_col.str.contains(value_str, literal=True)
                                                elif operator == "Não contém": expr = ~polars_col.str.contains(value_str, literal=True)
                                                elif operator == "Começa com": expr = polars_col.str.starts_with(value_str)
                                                elif operator == "Termina com": expr = polars_col.str.ends_with(value_str)
                                        
                                        if expr is not None:
                                            target_list.append(expr)

                                    except Exception as e_filter:
                                        self.log_message.emit(f"Não foi possível aplicar a regra de filtro '{col_name} {operator} {value}': {e_filter}", LogLevel.WARNING)
                                
                                # 2. Construir a expressão final para esta coluna
                                col_final_expr = None
                                
                                # Combinar todas as expressões de inclusão com OU (OR)
                                final_inclusion_expr = pl.any_horizontal(inclusion_exprs) if len(inclusion_exprs) > 1 else (inclusion_exprs[0] if inclusion_exprs else None)
                                
                                # Combinar todas as expressões de exclusão com E (AND)
                                final_exclusion_expr = pl.all_horizontal(exclusion_exprs) if len(exclusion_exprs) > 1 else (exclusion_exprs[0] if exclusion_exprs else None)

                                # Juntar inclusão e exclusão com E (AND)
                                if final_inclusion_expr is not None and final_exclusion_expr is not None:
                                    col_final_expr = final_inclusion_expr & final_exclusion_expr
                                elif final_inclusion_expr is not None:
                                    col_final_expr = final_inclusion_expr
                                elif final_exclusion_expr is not None:
                                    col_final_expr = final_exclusion_expr

                                if col_final_expr is not None:
                                    final_expressions_to_and.append(col_final_expr)

                            # 3. Aplicar os filtros finais combinados com E (AND)
                            if final_expressions_to_and:
                                rows_before = df_filtered.height
                                df_filtered = df_filtered.filter(final_expressions_to_and)
                                rows_after = df_filtered.height
                                self.log_message.emit(f"Filtro aplicado em {current_item_description}. Linhas restantes: {rows_after} de {rows_before}.", LogLevel.INFO)
                        
                        # --- Fim do Bloco de Filtros ---

                        # --- 3. Adicionar Coluna de Origem --- 
                        file_name_only = os.path.basename(file_path)
                        source_name = f"{file_name_only} ({sheet_name})" if sheet_name else file_name_only
                        
                        df_final_for_list = df_filtered.with_columns( # <-- Usa df_filtered
                            pl.lit(source_name).alias("Origem")
                        )

                        all_dataframes_processed.append(df_final_for_list)

                    except Exception as e:
                        self.log_message.emit(f"Erro ao processar (ler/mapear/tipar) {current_item_description}: {e}", LogLevel.ERROR)
                    
                    processed_items += 1
                    progress = int((processed_items / total_items) * 100) if total_items > 0 else 0
                    self.progress_updated.emit(progress)

                if not self.is_running: break 

            if not self.is_running:
                 self.log_message.emit("Consolidação cancelada.", LogLevel.WARNING)
                 self.finished.emit(False, "Cancelado"); return

            if not all_dataframes_processed:
                self.log_message.emit("Nenhum dado após processamento.", LogLevel.WARNING)
                self.finished.emit(False, "Nenhum dado processado."); return

            # --- Harmonização de Tipos (Pós-Tipagem do Usuário e Mapeamento) ---
            self.log_message.emit("Harmonizando tipos (2ª passagem) entre arquivos processados...", LogLevel.INFO)
            
            # 1. Coletar todos os tipos para cada nome de coluna final único
            #    em todos os DataFrames processados.
            column_all_types_globally = {} # {final_col_name: set_of_dtypes}
            for df_processed in all_dataframes_processed:
                 for col_name, dtype in df_processed.schema.items():
                     if col_name not in column_all_types_globally:
                         column_all_types_globally[col_name] = set()
                     column_all_types_globally[col_name].add(dtype)
            
            # 2. Determinar o tipo alvo para cada coluna globalmente
            global_target_types = {} # {final_col_name: target_polars_type}
            for final_col_name, dtypes_set in column_all_types_globally.items():
                is_int_present = any(t.is_integer() for t in dtypes_set)
                is_float_present = any(t.is_float() for t in dtypes_set)
                is_string_present = any(t == pl.String or t == pl.Utf8 for t in dtypes_set)
                is_temporal_present = any(t.is_temporal() for t in dtypes_set)
                is_boolean_present = any(t == pl.Boolean for t in dtypes_set)
                is_null_present = any(t == pl.Null for t in dtypes_set) # Null type

                target_type_for_col = None

                # Regra de Prioridade para determinar o tipo alvo:
                if is_string_present: # Se String estiver presente, tudo vira String
                    target_type_for_col = pl.String
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global String (devido à presença de String).", LogLevel.INFO)
                elif is_temporal_present and (is_int_present or is_float_present or is_boolean_present): # Temporal com outros não-string -> String
                    target_type_for_col = pl.String
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global String (conflito Temporal com Numérico/Booleano).", LogLevel.INFO)
                elif is_boolean_present and (is_int_present or is_float_present): # Booleano com Numérico -> String
                    target_type_for_col = pl.String
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global String (conflito Booleano com Numérico).", LogLevel.INFO)
                elif is_float_present: # Se Float estiver presente (e não String), tudo vira Float
                    target_type_for_col = pl.Float64
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global Decimal (Float) (devido à presença de Float ou Int+Float).", LogLevel.INFO)
                elif is_int_present: # Se apenas Int (e talvez Null, Boolean que pode ser Int)
                    target_type_for_col = pl.Int64 # Se só há Int e Null, pode ser Int. Se Booleano, pode ser Int.
                    # Se Booleano estiver presente e quisermos ser mais específicos, poderíamos ter mais regras
                    # Mas Int64 pode acomodar Booleanos como 0/1 se o cast funcionar.
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global Inteiro.", LogLevel.INFO)
                elif is_temporal_present: # Apenas Temporal (e talvez Null)
                    # Se houver múltiplos tipos temporais (Date, Datetime, Duration), escolher o mais geral (Datetime?) ou String.
                    # Por simplicidade, se só Temporal, manter (o primeiro que encontrar, ou o mais comum).
                    # Para ser seguro, se houver vários tipos temporais, converter para String ou o tipo mais abrangente.
                    # Esta parte pode precisar de mais refinamento se você tiver mistura de Date, Datetime, etc.
                    # Vamos pegar o primeiro tipo temporal encontrado como exemplo, ou default para pl.Date.
                    first_temporal_type = next((t for t in dtypes_set if t.is_temporal()), pl.Date)
                    target_type_for_col = first_temporal_type
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global {first_temporal_type} (apenas Temporal).", LogLevel.INFO)
                elif is_boolean_present: # Apenas Booleano (e talvez Null)
                    target_type_for_col = pl.Boolean
                    self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global Booleano.", LogLevel.INFO)
                # Se for apenas Null, ou tipos não cobertos, ele permanecerá None, e o cast não será aplicado abaixo
                # ou podemos definir um default como String.
                elif is_null_present and len(dtypes_set) == 1: # Apenas Null type
                     # Deixar como está por enquanto, o concat pode lidar com coluna toda Null.
                     # Ou, se quisermos ser proativos:
                     # target_type_for_col = pl.String
                     # self.log_message.emit(f"Coluna '{final_col_name}': Tipo alvo global String (era apenas Null).", LogLevel.INFO)
                     pass # Não define target_type, o cast não será aplicado para esta coluna se ela for só Null

                if target_type_for_col:
                    global_target_types[final_col_name] = target_type_for_col
            
            # 3. Aplicar o tipo alvo global a cada DataFrame
            harmonized_dataframes_final_pass = [] 
            for df_to_harmonize in all_dataframes_processed:
                 df_modified_this_pass = df_to_harmonize
                 
                 expressions_to_apply = []
                 for col_name_in_df in df_to_harmonize.columns:
                     target_type = global_target_types.get(col_name_in_df)
                     current_type = df_to_harmonize.schema.get(col_name_in_df)

                     if target_type and current_type != target_type:
                         # Só aplicar cast se o tipo atual for diferente do alvo
                         expressions_to_apply.append(pl.col(col_name_in_df).cast(target_type, strict=False).alias(col_name_in_df))
                         self.log_message.emit(f"Aplicando tipo alvo '{target_type}' à coluna '{col_name_in_df}' (era '{current_type}').", LogLevel.INFO)
                     else:
                         # Manter a coluna como está (ou porque não há tipo alvo ou já é o tipo alvo)
                         expressions_to_apply.append(pl.col(col_name_in_df))
                 
                 if expressions_to_apply: # Se houver colunas no DF (sempre deve haver se chegou aqui)
                     df_modified_this_pass = df_to_harmonize.select(expressions_to_apply)
                 
                 harmonized_dataframes_final_pass.append(df_modified_this_pass)
            
            final_dataframes_to_concat = harmonized_dataframes_final_pass
            # --- Fim Harmonização (2ª passagem) ---
            # dataframes_final_pass_no_null_type = []
            # for df in dataframes_final_pass_no_null_type:
            #     expression_for_null_cast = []
            #     needs_null_type_casting = False
            #     for col_name in df.columns:
            #         if df[col_name].dtype == pl.Null:
            #             expression_for_null_cast.append(pl.col(col_name).cast(pl.String).alias(col_name))
            #             needs_null_type_casting = True
            #         else:
            #             expression_for_null_cast.append(pl.col(col_name))
            #     if needs_null_type_casting and expression_for_null_cast:
            #         dataframes_final_pass_no_null_type.append(df.select(expression_for_null_cast))
            #     else:
            #         final_dataframes_to_concat = dataframes_final_pass_no_null_type
            self.log_message.emit("Concatenando dados processados...", LogLevel.INFO)
            try:
                consolidated_df = pl.concat(final_dataframes_to_concat, how="diagonal")
                # --- Reordenar Coluna "Origem" para o Final ---
                if "Origem" in consolidated_df.columns:
                    # Pega todas as colunas, exceto "Origem"
                    all_other_columns = [col for col in consolidated_df.columns if col != "Origem"]
                    # Cria a nova ordem com "Origem" no final
                    new_column_order = all_other_columns + ["Origem"]
                    # Seleciona as colunas na nova ordem
                    consolidated_df = consolidated_df.select(new_column_order)
                
                # --- Remoção de duplicatas ---
                removed_duplicates_df = None
                key_columns = self.duplicates_config.get("key_columns", [])
                generate_report = self.duplicates_config.get("generate_report", False)
                if key_columns:
                    rows_before = consolidated_df.height
                    self.log_message.emit(f"Removendo duplicatas com base nas chaves: {', '.join(key_columns)}...", LogLevel.INFO)
                    df_with_index = consolidated_df.with_row_index("__temp_index__")
                    unique_rows = df_with_index.unique(subset = key_columns, keep = 'first')
                    if generate_report:
                        removed_duplicates_df = df_with_index.join(unique_rows, on = "__temp_index__", how = "anti").drop("__temp_index__")
                    consolidated_df = unique_rows.drop("__temp_index__")
                    rows_after = consolidated_df.height
                    self.log_message.emit(f"{rows_before - rows_after} linhas duplicadas foram removidas. Linhas restantes: {rows_after}", LogLevel.SUCCESS)
                    if removed_duplicates_df is not None and not removed_duplicates_df.is_empty():
                        self.log_message.emit(f"Uma aba com as {removed_duplicates_df.height} linhas removidas será gerada.", LogLevel.INFO)
                
                # --- Aplicar Regras da Tabela de Resumo (Pivot) ---
                pivot_df = None # DataFrame para a tabela de resumo
                if self.pivot_rules and self.pivot_rules.get("group_by") and self.pivot_rules.get("aggregations"):
                    self.log_message.emit("Criando Tabela de Resumo (em memória)...", LogLevel.INFO)
                    try:
                        group_by_cols = self.pivot_rules['group_by']
                        aggregations = self.pivot_rules['aggregations']

                        op_map = {
                            "Soma": pl.sum, "Média": pl.mean, "Contagem": pl.count,
                            "Mínimo": pl.min, "Máximo": pl.max,
                            "Contagem Única": lambda col: pl.col(col).n_unique()
                        }

                        agg_expressions = []
                        for rule in aggregations:
                            col_name = rule['column']
                            op_str = rule['operation']
                            
                            if col_name not in consolidated_df.columns:
                                self.log_message.emit(f"Coluna '{col_name}' da regra de resumo não encontrada. Pulando.", LogLevel.WARNING)
                                continue
                            
                            if op_str in op_map:
                                polars_func = op_map[op_str]
                                new_col_name = f"{col_name}_{op_str.replace(' ', '_')}"
                                agg_expressions.append(polars_func(col_name).alias(new_col_name))
                        
                        if agg_expressions:
                            pivot_df = consolidated_df.group_by(group_by_cols).agg(agg_expressions).sort(group_by_cols)
                            self.log_message.emit("Tabela de resumo criada com sucesso.", LogLevel.SUCCESS)

                    except Exception as e_pivot:
                        self.log_message.emit(f"Erro ao criar tabela de resumo: {e_pivot}. O resultado do resumo não será salvo.", LogLevel.ERROR)
                        pivot_df = None
                # --- FIM DO BLOCO DE PIVOT --
            except Exception as e: 
                 self.log_message.emit(f"Erro concatenação final: {e}", LogLevel.ERROR)
                 self.finished.emit(False, f"Erro concatenação: {e}"); return
            if self.output_format == "XLSX":
                illegal_xml_chars_re = r"[\u0000-\u0008\u000B\u000C\u000E-\u001F]"
                # Sanitiza ambos os dataframes
                consolidated_df = consolidated_df.with_columns(
                    pl.col(pl.String).str.replace_all(illegal_xml_chars_re, "")
                )
                if pivot_df is not None:
                    pivot_df = pivot_df.with_columns(
                        pl.col(pl.String).str.replace_all(illegal_xml_chars_re, "")
                    )
            
            self.log_message.emit(f"Salvando: {self.output_path}", LogLevel.INFO)
            only_pivot = self.pivot_rules.get("only_pivot", False)

            if self.output_format == "XLSX":
                try:
                    import xlsxwriter
                    workbook = xlsxwriter.Workbook(self.output_path, {'use_zip64': True})
                    
                    # Formatos
                    header_format = workbook.add_format({'font_name': 'Aptos', 'bold': True, 'font_color': 'white', 'bg_color': '#000000', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
                    data_format = workbook.add_format({'font_name': 'Aptos'})
                    group_by_header_format = workbook.add_format({'font_name': 'Aptos', 'bold': True, 'font_color': 'white', 'bg_color': '#000000', 'border': 1, 'align': 'center', 'valign': 'vcenter'})

                    total_rows_to_write = consolidated_df.height + (pivot_df.height if pivot_df is not None else 0)
                    total_rows_written = 0

                    # 1. Escrever a Tabela de Resumo (pivot_df), se existir
                    if pivot_df is not None:
                        self.log_message.emit("Escrevendo aba 'Tabela_Resumo'...", LogLevel.INFO)
                        worksheet_pivot = workbook.add_worksheet("Tabela_Resumo")
                        worksheet_pivot.freeze_panes('A2')
                        worksheet_pivot.set_zoom(70)
                        worksheet_pivot.hide_gridlines(2)
                        
                        group_by_cols = self.pivot_rules.get('group_by', [])
                        for col_idx, col_name in enumerate(pivot_df.columns):
                            fmt = group_by_header_format if col_name in group_by_cols else header_format
                            worksheet_pivot.write(0, col_idx, col_name, fmt)
                        
                        max_lengths_pivot = {col: max(len(str(col)), pivot_df.select(pl.col(col).cast(pl.String).str.len_chars().max()).item() or 1) for col in pivot_df.columns}
                        for col_idx, col_name in enumerate(pivot_df.columns):
                            width = min(max_lengths_pivot.get(col_name, len(col_name)) + 2, 60)
                            worksheet_pivot.set_column(col_idx, col_idx, width, data_format)
                        
                        for r_idx, row_tuple in enumerate(pivot_df.iter_rows(), start=1):
                            worksheet_pivot.write_row(r_idx, 0, row_tuple)
                            total_rows_written += 1
                    if not only_pivot:
                        # 2. Escrever os Dados Consolidados 
                        self.log_message.emit("Escrevendo aba(s) de 'Dados_Consolidados'...", LogLevel.INFO)
                        max_rows_per_sheet = 1_048_570
                        dfs_to_write = []
                        if consolidated_df.height > max_rows_per_sheet:
                            num_chunks = (consolidated_df.height + max_rows_per_sheet - 1) // max_rows_per_sheet
                            for i in range(num_chunks):
                                df_chunk = consolidated_df.slice(i * max_rows_per_sheet, max_rows_per_sheet)
                                dfs_to_write.append((f"Dados_Consolidados_{i+1}", df_chunk))
                        else:
                            dfs_to_write.append(("Dados_Consolidados", consolidated_df))

                        max_lengths_consolidated = {col: max(len(str(col)), consolidated_df.select(pl.col(col).cast(pl.String).str.len_chars().max()).item() or 1) for col in consolidated_df.columns}
                        if removed_duplicates_df is not None and not removed_duplicates_df.is_empty():
                            self.log_message.emit("Escrevendo aba 'Duplicatas_Removidas'...", LogLevel.INFO)
                            duplicates_header_format = workbook.add_format({'font_name': 'Aptos', 'bold': True, 'font_color': 'white', 'bg_color': '#C00000', 'border': 1, 'align': 'center', 'valign': 'vcenter'}) # Cabeçalho vermelho

                            # A lógica de chunking é importante aqui também
                            max_rows_per_sheet = 1_048_570
                            duplicates_to_write = []
                            if removed_duplicates_df.height > max_rows_per_sheet:
                                num_chunks = (removed_duplicates_df.height + max_rows_per_sheet - 1) // max_rows_per_sheet
                                for i in range(num_chunks):
                                    duplicates_to_write.append((f"Duplicatas_Removidas_{i+1}", removed_duplicates_df.slice(i * max_rows_per_sheet, max_rows_per_sheet)))
                            else:
                                duplicates_to_write.append(("Duplicatas_Removidas", removed_duplicates_df))
                            
                            max_lengths_duplicates = {col: max(len(str(col)), removed_duplicates_df.select(pl.col(col).cast(pl.String).str.len_chars().max()).item() or 1) for col in removed_duplicates_df.columns}

                            for sheet_name, df_chunk in duplicates_to_write:
                                worksheet = workbook.add_worksheet(sheet_name)
                                worksheet.freeze_panes('A2')
                                worksheet.set_zoom(70)
                                worksheet.hide_gridlines(2)
                                worksheet.write_row('A1', df_chunk.columns, duplicates_header_format)
                                worksheet.autofilter(0, 0, df_chunk.height, df_chunk.width - 1)
                                for col_idx, col_name in enumerate(df_chunk.columns):
                                    worksheet.set_column(col_idx, col_idx, min(max_lengths_duplicates.get(col_name, len(col_name)) + 2, 60), data_format)
                                for r_idx, row_tuple in enumerate(df_chunk.iter_rows(), start=1):
                                    worksheet.write_row(r_idx, 0, row_tuple)
                                    total_rows_written += 1
                                    if total_rows_written % 5000 == 0:
                                        self.progress_text_updated.emit(f"Escrevendo linha {total_rows_written:,} de {total_rows_to_write:,}...")
                                
                            
                            # max_lengths_duplicates = {col: max(len(str(col)), removed_duplicates_df.select(pl.col(col).cast(pl.String).str.len_chars().max()).item() or 1) for col in removed_duplicates_df.columns}
                        for sheet_name, df_chunk in dfs_to_write:
                            worksheet = workbook.add_worksheet(sheet_name)
                            worksheet.freeze_panes('A2')
                            worksheet.set_zoom(70)
                            worksheet.hide_gridlines(2)
                            worksheet.write_row('A1', df_chunk.columns, header_format)
                            worksheet.autofilter(0, 0, df_chunk.height, df_chunk.width - 1)
                            for col_idx, col_name in enumerate(df_chunk.columns):
                                width = min(max_lengths_consolidated.get(col_name, len(col_name)) + 2, 60)
                                worksheet.set_column(col_idx, col_idx, width, data_format)
                            
                            for r_idx, row_tuple in enumerate(df_chunk.iter_rows(), start=1):
                                worksheet.write_row(r_idx, 0, row_tuple)
                                total_rows_written += 1
                                if total_rows_written % 5000 == 0:
                                    self.progress_text_updated.emit(f"Escrevendo linha {total_rows_written:,} de {total_rows_to_write:,}...")
                        
                        self.progress_text_updated.emit(f"Finalizando escrita de {total_rows_to_write:,} linhas...")
                    workbook.close()

                except Exception as e_save_excel:
                    self.log_message.emit(f"Erro ao salvar arquivo Excel com XlsxWriter: {e_save_excel}", LogLevel.ERROR)
                    self.finished.emit(False, f"Erro ao salvar Excel: {e_save_excel}")
                    return

            elif self.output_format in ["CSV", "Parquet"]:
                df_to_save = pivot_df if pivot_df is not None else consolidated_df
                if pivot_df is not None:
                    self.log_message.emit(f"Salvando resultado da Tabela de Resumo em {self.output_format}.", LogLevel.INFO)
                
                if self.output_format == "CSV":
                    df_to_save.write_csv(self.output_path, separator='|')
                elif self.output_format == "Parquet":
                    df_to_save.write_parquet(self.output_path, compression='zstd')

            self.progress_updated.emit(100)
            self.log_message.emit(f"Concluído! Salvo em: {self.output_path}", LogLevel.SUCCESS)
            self.finished.emit(True, f"Salvo em: {self.output_path}")

        except Exception as e:
            self.log_message.emit(f"Erro inesperado consolidação: {e}", LogLevel.ERROR)
            self.finished.emit(False, f"Erro: {e}")

    def stop(self): # stop() permanece o mesmo
        self.is_running = False
        self.log_message.emit("Tentativa de parada da consolidação solicitada...", LogLevel.INFO)

class SheetLoadingWorker(QThread):
    finished = Signal(str, list, str)  # file_path, sheet_names_list, error_message_or_None

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path
        self.is_running = True

    def run(self):
        if not self.is_running:
            self.finished.emit(self.file_path, [], "Cancelado antes de iniciar.")
            return

        sheet_names_from_file = []
        error_message = None
        try:
            if not self.is_running: # Checar novamente
                raise InterruptedError("Carregamento de abas cancelado.")

            if self.file_path.lower().endswith(".xlsx"):
                # Usar read_only=True para performance, data_only=True para evitar carregar fórmulas complexas se não necessário
                workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
                sheet_names_from_file = workbook.sheetnames
                workbook.close() # É bom fechar o workbook
            elif self.file_path.lower().endswith(".xls"):
                workbook = xlrd.open_workbook(self.file_path, on_demand=True)
                sheet_names_from_file = workbook.sheet_names()
            # Não precisa de else, pois a thread só é chamada para .xlsx/.xls
        except InterruptedError as ie:
            error_message = str(ie)
        except Exception as e:
            error_message = f"Erro ao ler abas do arquivo {os.path.basename(self.file_path)}: {e}"
        
        if self.is_running: # Só emite se não foi cancelado durante a operação
            self.finished.emit(self.file_path, sheet_names_from_file, error_message)

    def stop(self):
        self.is_running = False

class SheetAnalysisWorker(QThread):
    """
    Worker para analisar todos os arquivos Excel em uma lista e retornar
    um cache com todas as abas por arquivo e um conjunto de nomes de abas únicos.
    """
    # Sinal emite: {cache_de_abas}, {abas_unicas}, mensagem_de_erro
    finished = Signal(dict, set, str)

    def __init__(self, excel_files_paths):
        super().__init__()
        self.excel_files_paths = excel_files_paths
        self.is_running = True

    def run(self):
        all_sheets_cache = {} # Ex: {'caminho/arquivo1.xlsx': ['Aba1', 'Aba2'], ...}
        unique_sheet_names = set()
        error_message = None

        try:
            for file_path in self.excel_files_paths:
                if not self.is_running:
                    raise InterruptedError("Análise de abas cancelada.")
                
                try:
                    sheet_names = []
                    # Usamos openpyxl diretamente pois é otimizado para apenas ler nomes de abas
                    if file_path.lower().endswith(".xlsx"):
                        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        sheet_names = workbook.sheetnames
                        workbook.close()
                    elif file_path.lower().endswith(".xls"):
                        workbook = xlrd.open_workbook(file_path, on_demand=True)
                        sheet_names = workbook.sheet_names()
                    
                    if sheet_names:
                        all_sheets_cache[file_path] = sheet_names
                        unique_sheet_names.update(sheet_names)

                except Exception as e:
                    # Loga um erro para um arquivo específico mas continua o processo
                    print(f"AVISO: Não foi possível ler as abas de '{os.path.basename(file_path)}'. Erro: {e}")
            
            if self.is_running:
                self.finished.emit(all_sheets_cache, unique_sheet_names, None)

        except InterruptedError as ie:
            error_message = str(ie)
        except Exception as e:
            error_message = f"Um erro inesperado ocorreu durante a análise de abas: {e}"
        
        if error_message:
            self.finished.emit({}, set(), error_message)

    def stop(self):
        self.is_running = False

class HeaderAnalysisWorker(QThread):
    '''Worker para os cabeçalhos'''
    finished = Signal(list, object)
    progress_log = Signal(str, LogLevel)

    def __init__(self, files_and_sheets_config, delimiter):
        super().__init__()
        self.files_and_sheets_config = files_and_sheets_config
        self.delimiter = delimiter
        self.is_running = True

    def _get_series_profile(self, series: pl.Series):
        non_null_series = series.filter(series.is_not_null() & (series.str.strip_chars() != ""))
        if non_null_series.is_empty():
            return {"dtype": pl.String, "null_ratio": series.is_null().mean()}
        try:
            if non_null_series.cast(pl.Int64, strict=True).is_not_null().all():
                return {"dtype": pl.Int64, "null_ratio": series.is_null().mean()}
        except (Exception, pl.exceptions.PanicException): pass
        try:
            if non_null_series.cast(pl.Float64, strict=True).is_not_null().all():
                return {"dtype": pl.Float64, "null_ratio": series.is_null().mean()}
        except (Exception, pl.exceptions.PanicException): pass
        try:
            if non_null_series.str.to_datetime(strict=True, exact=False, cache=False).is_not_null().all():
                return {"dtype": pl.Datetime, "null_ratio": series.is_null().mean()}
        except (Exception, pl.exceptions.PanicException): pass
        return {"dtype": pl.String, "null_ratio": series.is_null().mean()}

    def run(self):
        if not self.is_running:
            self.finished.emit([], InterruptedError("Análise cancelada."))
            return

        all_column_fingerprints = []
        n_sample_rows, n_preread_rows = 200, 20

        try:
            for file_path, selected_sheets in self.files_and_sheets_config:
                if not self.is_running: raise InterruptedError("Análise cancelada.")

                self.progress_log.emit(f"Analisando: {os.path.basename(file_path)}...", LogLevel.INFO)
                
                sheets_to_iterate = selected_sheets if selected_sheets is not None else [None]
                for sheet_name in sheets_to_iterate:
                    if not self.is_running: raise InterruptedError("Análise cancelada.")
                    try:
                        pre_read_df = None
                        if file_path.lower().endswith((".csv", ".txt")):
                            pre_read_df = pl.read_csv(source=file_path, has_header=False, n_rows=n_preread_rows, separator=self.delimiter, encoding='latin-1', ignore_errors=True, infer_schema = False, truncate_ragged_lines = True, quote_char = None)
                        elif file_path.lower().endswith((".xlsx", ".xls")):
                            pre_read_df = pl.read_excel(source=file_path, sheet_name=sheet_name, has_header = False, infer_schema_length = 0).head(n_preread_rows)
                        
                        if pre_read_df is None or pre_read_df.is_empty(): continue
                        
                        header_row_index = _find_header_row_index(pre_read_df, n_preread_rows)
                        header_names_raw = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(pre_read_df.row(header_row_index))]
                        header_names = _make_headers_unique(header_names_raw)
                        data_rows_df = pre_read_df.slice(offset=header_row_index + 1).head(n_sample_rows)
                        
                        if data_rows_df.is_empty(): continue
                        
                        rename_mapping = {old_name: new_name for old_name, new_name in zip(data_rows_df.columns, header_names)}
                        sample_df = data_rows_df.rename(rename_mapping)

                        for col_name in sample_df.columns:
                            try: # <-- INÍCIO DO BLOCO DE BLINDAGEM
                                series = sample_df[col_name]
                                profile = self._get_series_profile(series)
                                fingerprint = {
                                    "source_tuple": (col_name, file_path, sheet_name),
                                    "normalized_name": _normalize_header_name(col_name),
                                    "dtype": profile["dtype"],
                                    "null_ratio": profile["null_ratio"],
                                }
                            except Exception as e_profile: # <-- CAPTURA O "PANIC"
                                # Se a análise da coluna falhar, cria um perfil "seguro"
                                self.progress_log.emit(f"Falha ao analisar coluna '{col_name}' em '{os.path.basename(file_path)}'. Tratando como texto. Erro: {e_profile}", LogLevel.WARNING)
                                fingerprint = {
                                    "source_tuple": (col_name, file_path, sheet_name),
                                    "normalized_name": _normalize_header_name(col_name),
                                    "dtype": pl.String, # Tipo de dado seguro
                                    "null_ratio": 0.0,
                                }
                            all_column_fingerprints.append(fingerprint)
                    except Exception:
                        continue

            # --- ALGORITMO DE AGRUPAMENTO DEFINITIVO ---
            groups_by_name = defaultdict(list)
            for fp in all_column_fingerprints:
                groups_by_name[fp["normalized_name"]].append(fp)
            
            final_groups = []
            for normalized_name, fp_list in groups_by_name.items():
                if len(fp_list) == 1:
                    final_groups.append([fp_list[0]["source_tuple"]])
                    continue

                sub_groups_by_type = defaultdict(list)
                for fp in fp_list:
                    type_key = "numeric" if fp["dtype"].is_numeric() else str(fp["dtype"])
                    sub_groups_by_type[type_key].append(fp["source_tuple"])
                
                for type_key, source_tuples in sub_groups_by_type.items():
                    final_groups.append(source_tuples)

            if self.is_running:
                self.finished.emit(final_groups, None)
        except Exception as e:
            self.finished.emit([], e)

    def stop(self):
        self.is_running = False
